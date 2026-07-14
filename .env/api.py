# api.py - Sistema de Escala de Limpeza - Backend Completo
# VERSÃO COM SISTEMA COMPLETO DE PERMISSÕES RBAC
# COM BANCO DE DADOS SEPARADO PARA ADMINS

import hashlib
from werkzeug.security import generate_password_hash, check_password_hash
import os
import json
import sqlite3
import random
import secrets
import csv
import io
import re
import hashlib
import hmac
import time
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_file, send_from_directory, session, g
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import bleach
import magic
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from PIL import Image
import mimetypes

# ============================================
# CONFIGURAÇÃO INICIAL
# ============================================

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

app.config['SECRET_KEY'] = secrets.token_hex(32)
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'None'
app.config['SESSION_COOKIE_DOMAIN'] = '.pythonanywhere.com'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# ============================================
# RATE LIMITING - Proteção contra força bruta
# ============================================

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["500 per day", "100 per hour"],
    storage_uri="memory://"
)

# ============================================
# CORS CONFIGURADO CORRETAMENTE
# ============================================

ALLOWED_ORIGINS = [
    'https://escalalimpeza.vercel.app',
    'https://adminescala.vercel.app',
    'http://localhost:3000',
    'http://localhost:5000',
    "http://127.0.0.1:5500"
]

CORS(app, resources={
    r"/api/*": {
        "origins": ALLOWED_ORIGINS,
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "expose_headers": ["Content-Type"],
        "supports_credentials": True,
        "max_age": 3600
    },
    r"/uploads/*": {
        "origins": ALLOWED_ORIGINS,
        "methods": ["GET"],
        "allow_headers": ["*"]
    },
    r"/thumbnails/*": {
        "origins": ALLOWED_ORIGINS,
        "methods": ["GET"],
        "allow_headers": ["*"]
    }
})

# No início do arquivo api.py, após definir ALLOWED_ORIGINS, adicione:

@app.after_request
def add_cors_headers(response):
    """Adiciona headers CORS em todas as respostas"""
    origin = request.headers.get('Origin')
    if origin in ALLOWED_ORIGINS:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    return response

@app.route('/api/mentoria/ciclos', methods=['OPTIONS'])
@app.route('/api/mentoria/formularios', methods=['OPTIONS'])
@app.route('/api/mentoria/alunos', methods=['OPTIONS'])
@app.route('/api/mentoria/mentores', methods=['OPTIONS'])
@app.route('/api/mentoria/relacoes', methods=['OPTIONS'])
@app.route('/api/mentoria/turmas', methods=['OPTIONS'])
@app.route('/api/mentoria/status', methods=['OPTIONS'])
@app.route('/api/mentoria/importar', methods=['OPTIONS'])
def handle_options():
    """Manipula requisições OPTIONS para CORS"""
    return '', 200
# ============================================
# HEADERS DE SEGURANÇA
# ============================================

@app.after_request
def add_security_headers(response):
    """Adiciona headers de segurança em todas as respostas"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self'; style-src 'self'; img-src 'self' data:;"
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # Permite geolocalização no próprio domínio (necessário para as páginas de presença)
    response.headers['Permissions-Policy'] = "geolocation=(self), microphone=(), camera=()"
    return response

# ============================================
# CONFIGURAÇÕES DO SISTEMA
# ============================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, 'db.sqlite3')
ADMINS_DB = os.path.join(BASE_DIR, 'admins.db')
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
THUMBNAIL_DIR = os.path.join(BASE_DIR, 'thumbnails')
SECRET_KEY = secrets.token_hex(32)

THUMBNAIL_SIZE = (150, 150)
THUMBNAIL_QUALITY = 85

# ============================================
# SISTEMA DE PERMISSÕES E TIPOS DE ADMIN
# ============================================

# Definição das turmas disponíveis
TURMAS = {
    'formare': 'Formare 2026',
    'aprender_terca': 'Aprender A+ (Terça-Feira)',
    'aprender_quarta': 'Aprender A+ (Segunda-Feira)',
    'aprender_quinta': 'Aprender A+ (Quinta-Feira)',
    'informatica_basica_curitiba': 'Informática Básica (Curitiba)',
    'ingles_basico_curitiba': 'Inglês Básico (Curitiba)',
    'informatica_robotica_curitiba': 'Informática II - Robótica (Curitiba)',
    'ingles_basico_joinville': 'Inglês Básico (Joinville)',
    'ingles_basico_pomerode': 'Inglês Básico (Pomerode)',
    'informatica_basica_joinville': 'Informática Básica (Joinville)'
}

# Mapeamento reverso para facilitar buscas
TURMAS_REVERSE = {v: k for k, v in TURMAS.items()}

# ============================================
# DEFINIÇÃO DE PERMISSÕES POR TIPO DE ADMIN
# ============================================

PERMISSOES_ADMIN_GLOBAL = [
    'presenca', 'upload', 'avisos', 'backup', 'escalas', 'alunos', 'historico', 'relatorios', 'todas_turmas'
]

PERMISSOES_ADMIN_FORMARE = [
    'presenca_formare_only', 'visualizar_todas_turmas', 'upload', 'avisos', 'backup', 'escalas', 'alunos', 'historico', 'relatorios'
]

PERMISSOES_ADMIN_TURMA = [
    'presenca_apenas_propria_turma',
    'relatorios'   # ← ADICIONE ESTA LINHA
]

# ============================================
# BANCO DE DADOS DE ADMINS (SEPARADO)
# ============================================

def init_admins_db():
    """Inicializa o banco de dados de administradores"""
    conn = sqlite3.connect(ADMINS_DB)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            admin_type TEXT NOT NULL,
            turma_permitida TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS admin_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            action TEXT NOT NULL,
            details TEXT,
            ip TEXT,
            user_agent TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS blocked_ips (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ip TEXT UNIQUE NOT NULL,
            reason TEXT,
            blocked_by TEXT,
            blocked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP
        )
    ''')

    conn.commit()
    conn.close()
    print("✅ Banco de dados de admins inicializado")

def get_admins_db():
    conn = sqlite3.connect(ADMINS_DB)
    conn.row_factory = sqlite3.Row
    return conn

def get_admin_by_username(username):
    """Busca um admin pelo username"""
    conn = get_admins_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM admins WHERE username = ?', (username,))
    admin = cursor.fetchone()
    conn.close()
    return admin

def get_all_admins():
    """Retorna todos os admins"""
    conn = get_admins_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, admin_type, turma_permitida, created_by, created_at FROM admins ORDER BY admin_type, username')
    admins = cursor.fetchall()
    conn.close()
    return admins

def create_admin(username, password, admin_type, turma_permitida=None, created_by='Sistema'):
    # Se for lista, converte para string separada por vírgula
    if isinstance(turma_permitida, list):
        turma_permitida = ','.join(turma_permitida)
    """Cria um novo administrador"""
    conn = get_admins_db()
    cursor = conn.cursor()

    password_hash = generate_password_hash(password)

    cursor.execute('''
        INSERT INTO admins (username, password_hash, admin_type, turma_permitida, created_by)
        VALUES (?, ?, ?, ?, ?)
    ''', (username, password_hash, admin_type, turma_permitida, created_by))

    conn.commit()
    admin_id = cursor.lastrowid
    conn.close()

    return admin_id

def update_admin_password(username, new_password):
    """Atualiza a senha de um admin"""
    conn = get_admins_db()
    cursor = conn.cursor()

    password_hash = generate_password_hash(new_password)

    cursor.execute('''
        UPDATE admins SET password_hash = ?, updated_at = CURRENT_TIMESTAMP
        WHERE username = ?
    ''', (password_hash, username))

    conn.commit()
    affected = cursor.rowcount
    conn.close()

    return affected > 0

def delete_admin(username):
    """Remove um administrador"""
    conn = get_admins_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM admins WHERE username = ?', (username,))
    affected = cursor.rowcount
    conn.commit()
    conn.close()
    return affected > 0

def log_admin_action_db(username, action, details=None, ip=None, user_agent=None):
    """Registra ação de admin no banco de dados de admins"""
    try:
        conn = get_admins_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO admin_logs (username, action, details, ip, user_agent)
            VALUES (?, ?, ?, ?, ?)
        ''', (username, action, details, ip, user_agent))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Erro ao registrar log de admin: {e}")

def add_blocked_ip(ip, reason, blocked_by='Sistema', expires_minutes=60):
    """Adiciona um IP à lista de bloqueio"""
    conn = get_admins_db()
    cursor = conn.cursor()
    expires_at = datetime.now() + timedelta(minutes=expires_minutes)

    cursor.execute('''
        INSERT OR REPLACE INTO blocked_ips (ip, reason, blocked_by, expires_at)
        VALUES (?, ?, ?, ?)
    ''', (ip, reason, blocked_by, expires_at))

    conn.commit()
    conn.close()

def remove_blocked_ip(ip):
    """Remove um IP da lista de bloqueio"""
    conn = get_admins_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM blocked_ips WHERE ip = ?', (ip,))
    affected = cursor.rowcount
    conn.commit()
    conn.close()
    return affected > 0

def is_ip_blocked(ip):
    """Verifica se um IP está bloqueado"""
    conn = get_admins_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM blocked_ips WHERE ip = ? AND (expires_at IS NULL OR expires_at > datetime('now'))
    ''', (ip,))
    blocked = cursor.fetchone()
    conn.close()
    return blocked is not None

def get_blocked_ips():
    """Retorna lista de IPs bloqueados"""
    conn = get_admins_db()
    cursor = conn.cursor()
    cursor.execute('SELECT ip, reason, blocked_by, blocked_at, expires_at FROM blocked_ips ORDER BY blocked_at DESC')
    ips = cursor.fetchall()
    conn.close()
    return ips

# ============================================
# FUNÇÕES DE PERMISSÃO (USANDO BANCO DE DADOS)
# ============================================

def get_admin_type(username):
    """Retorna o tipo de administrador baseado no banco de dados"""
    admin = get_admin_by_username(username)
    if admin:
        return admin['admin_type']
    return None

def get_admin_turmas_permitidas(username):
    admin = get_admin_by_username(username)
    admin_type = get_admin_type(username)  # <-- FALTOU ESTA LINHA!
    if admin_type == 'turma' and admin and admin['turma_permitida']:
        if ',' in admin['turma_permitida']:
            return admin['turma_permitida'].split(',')
        return [admin['turma_permitida']]
    return []

def get_admin_permissoes(username):
    """Retorna as permissões do administrador baseado no tipo"""
    admin_type = get_admin_type(username)

    if admin_type == 'global':
        return PERMISSOES_ADMIN_GLOBAL
    elif admin_type == 'formare':
        return PERMISSOES_ADMIN_FORMARE
    elif admin_type == 'turma':
        return PERMISSOES_ADMIN_TURMA
    return []

def can_visualize_turma(username, turma_nome):
    """Verifica se o admin pode visualizar uma turma"""
    admin_type = get_admin_type(username)
    turmas_permitidas = get_admin_turmas_permitidas(username)

    if admin_type in ['global', 'formare']:
        return True
    elif admin_type == 'turma':
        return turma_nome in turmas_permitidas
    return False

def can_perform_action(username, acao, turma_nome=None):
    """Verifica se o admin pode realizar uma ação específica"""
    admin_type = get_admin_type(username)
    permissoes = get_admin_permissoes(username)

    if admin_type == 'global':
        return True
    elif admin_type == 'formare':
        if acao == 'presenca':
            return turma_nome == TURMAS['formare']
        return acao in permissoes
    elif admin_type == 'turma':
        if acao == 'presenca':
            turmas_permitidas = get_admin_turmas_permitidas(username)
            return turma_nome in turmas_permitidas
        # ← ADICIONE ESTA LINHA:
        return acao in permissoes
    return False

def is_super_admin(username):
    """Verifica se é super admin (primeiro admin global criado)"""
    try:
        conn = get_admins_db()
        cursor = conn.cursor()

        # Buscar o primeiro admin global (super admin)
        cursor.execute('SELECT id FROM admins WHERE admin_type = "global" ORDER BY id LIMIT 1')
        first_result = cursor.fetchone()

        # Buscar o admin atual
        cursor.execute('SELECT id FROM admins WHERE username = ?', (username,))
        admin_result = cursor.fetchone()

        conn.close()

        if not first_result or not admin_result:
            return False

        # Extrair IDs independentemente do formato
        if isinstance(first_result, dict) or hasattr(first_result, 'keys'):
            first_id = first_result['id']
        else:
            first_id = first_result[0]

        if isinstance(admin_result, dict) or hasattr(admin_result, 'keys'):
            admin_id = admin_result['id']
        else:
            admin_id = admin_result[0]

        return first_id == admin_id

    except Exception as e:
        print(f"Erro em is_super_admin: {e}")
        return False
# ============================================
# ROTA DE LOGIN COM TOKEN
# ============================================

import jwt

JWT_SECRET = secrets.token_hex(32)
JWT_EXPIRATION = 8

# ============================================
# EDVS PERMITIDOS (acesso global read-only)
# ============================================

EDVS_PERMITIDOS = [
    {"edv": "92885075", "nome": "Marineide Maia"},
    {"edv": "92909207", "nome": "Erick Matheus"},
    {"edv": "19098",    "nome": "Larissa Amabilly"},
    {"edv": "92892918", "nome": "Eduardo Marcomini"},
    {"edv": "92896450", "nome": "Rayen Estefany Siqueira Negretti"},
]

# Rate limiting para tentativas de login
login_attempts = {}

# Data de início da primeira escala - 30 de março de 2026
DATA_INICIO_ESCALA = datetime(2026, 4, 6)

# Data limite para geração automática (31/12/2026)
DATA_LIMITE_AUTOMATICA = datetime(2026, 12, 31)

# ============================================
# FUNÇÕES DE SANITIZAÇÃO
# ============================================

ALLOWED_TAGS = ['strong', 'em', 'p', 'br', 'span', 'div', 'h1', 'h2', 'h3', 'h4', 'ul', 'li', 'a']
ALLOWED_ATTRIBUTES = {
    'a': ['href', 'title', 'target'],
    'span': ['class'],
    'div': ['class'],
}

def sanitizar_html(html_content):
    if not html_content:
        return ''
    return bleach.clean(html_content, tags=ALLOWED_TAGS, attributes=ALLOWED_ATTRIBUTES, strip=True)

def sanitizar_texto(texto):
    if not texto:
        return ''
    texto = re.sub(r'[<>]', '', texto)
    return texto.strip()

def validar_edv(edv):
    if not edv:
        return False
    return bool(re.match(r'^\d+$', edv))

def validar_nome(nome):
    if not nome:
        return False
    return bool(re.match(r'^[a-zA-ZáéíóúâêîôûãõçÁÉÍÓÚÂÊÎÔÛÃÕÇ\s\-\']+$', nome))

# ============================================
# FUNÇÕES PARA MINIATURAS
# ============================================

def get_file_icon(mime_type, filename):
    icons = {
        'image': '🖼️',
        'video': '🎬',
        'audio': '🎵',
        'pdf': '📄',
        'word': '📝',
        'excel': '📊',
        'powerpoint': '📽️',
        'text': '📃',
        'archive': '🗜️',
        'code': '💻',
        'default': '📁'
    }

    if mime_type.startswith('image/'):
        return icons['image']
    elif mime_type.startswith('video/'):
        return icons['video']
    elif mime_type.startswith('audio/'):
        return icons['audio']
    elif mime_type == 'application/pdf':
        return icons['pdf']
    elif 'word' in mime_type or mime_type == 'application/msword':
        return icons['word']
    elif 'excel' in mime_type or mime_type == 'application/vnd.ms-excel':
        return icons['excel']
    elif 'powerpoint' in mime_type:
        return icons['powerpoint']
    elif mime_type.startswith('text/'):
        return icons['text']
    elif mime_type in ['application/zip', 'application/x-rar-compressed', 'application/x-tar']:
        return icons['archive']
    elif mime_type in ['application/json', 'text/html', 'text/css', 'text/javascript']:
        return icons['code']
    else:
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        if ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
            return icons['image']
        elif ext in ['mp4', 'avi', 'mov', 'mkv']:
            return icons['video']
        elif ext in ['mp3', 'wav', 'ogg']:
            return icons['audio']
        elif ext == 'pdf':
            return icons['pdf']
        elif ext in ['doc', 'docx']:
            return icons['word']
        elif ext in ['xls', 'xlsx']:
            return icons['excel']
        elif ext in ['ppt', 'pptx']:
            return icons['powerpoint']
        elif ext in ['txt', 'md', 'rtf']:
            return icons['text']
        elif ext in ['zip', 'rar', '7z', 'tar', 'gz']:
            return icons['archive']
        elif ext in ['html', 'css', 'js', 'py', 'java', 'c', 'cpp']:
            return icons['code']
        else:
            return icons['default']

def generate_thumbnail(filepath, filename, mime_type):
    try:
        if not mime_type.startswith('image/'):
            return None

        if not os.path.exists(THUMBNAIL_DIR):
            os.makedirs(THUMBNAIL_DIR)

        thumbnail_name = f"thumb_{filename}"
        thumbnail_path = os.path.join(THUMBNAIL_DIR, thumbnail_name)

        if os.path.exists(thumbnail_path):
            if os.path.getmtime(thumbnail_path) >= os.path.getmtime(filepath):
                return f'/thumbnails/{thumbnail_name}'

        with Image.open(filepath) as img:
            if img.mode in ('RGBA', 'LA', 'P'):
                rgb_img = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                rgb_img.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = rgb_img

            img.thumbnail(THUMBNAIL_SIZE, Image.Resampling.LANCZOS)
            img.save(thumbnail_path, 'JPEG', quality=THUMBNAIL_QUALITY, optimize=True)

            return f'/thumbnails/{thumbnail_name}'

    except Exception as e:
        print(f"Erro ao gerar miniatura para {filename}: {e}")
        return None

def validate_file_path(base_dir, filename):
    try:
        filename = sanitizar_texto(filename)
        if not filename:
            return False

        if '..' in filename or filename.startswith('/') or filename.startswith('\\'):
            return False

        full_path = os.path.abspath(os.path.join(base_dir, filename))
        base_dir_abs = os.path.abspath(base_dir)

        return full_path.startswith(base_dir_abs)

    except Exception:
        return False

# ============================================
# BANCO DE DADOS PRINCIPAL
# ============================================

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def migrate_presenca_table():
    """Adiciona colunas de auditoria e localização na tabela presenca se não existirem"""
    conn = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute("PRAGMA table_info(presenca)")
        columns = [col[1] for col in cursor.fetchall()]

        if 'created_by' not in columns:
            cursor.execute("ALTER TABLE presenca ADD COLUMN created_by TEXT")
            print("✅ Coluna created_by adicionada")

        if 'updated_by' not in columns:
            cursor.execute("ALTER TABLE presenca ADD COLUMN updated_by TEXT")
            print("✅ Coluna updated_by adicionada")

        if 'created_lat' not in columns:
            cursor.execute("ALTER TABLE presenca ADD COLUMN created_lat REAL")
            print("✅ Coluna created_lat adicionada")

        if 'created_lng' not in columns:
            cursor.execute("ALTER TABLE presenca ADD COLUMN created_lng REAL")
            print("✅ Coluna created_lng adicionada")

        if 'created_loc_accuracy' not in columns:
            cursor.execute("ALTER TABLE presenca ADD COLUMN created_loc_accuracy REAL")
            print("✅ Coluna created_loc_accuracy adicionada")

        if 'updated_lat' not in columns:
            cursor.execute("ALTER TABLE presenca ADD COLUMN updated_lat REAL")
            print("✅ Coluna updated_lat adicionada")

        if 'updated_lng' not in columns:
            cursor.execute("ALTER TABLE presenca ADD COLUMN updated_lng REAL")
            print("✅ Coluna updated_lng adicionada")

        if 'updated_loc_accuracy' not in columns:
            cursor.execute("ALTER TABLE presenca ADD COLUMN updated_loc_accuracy REAL")
            print("✅ Coluna updated_loc_accuracy adicionada")

        cursor.execute("UPDATE presenca SET created_by = responsavel WHERE created_by IS NULL AND responsavel IS NOT NULL")
        cursor.execute("UPDATE presenca SET updated_by = NULL")

        conn.commit()
        print("✅ Migração da tabela presenca concluída")
    except Exception as e:
        print(f"Erro na migração: {e}")
    finally:
        conn.close()

def init_db():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS turmas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alunos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            edv TEXT NOT NULL,
            nome TEXT NOT NULL,
            turma_id INTEGER NOT NULL,
            is_admin INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(edv, turma_id),
            FOREIGN KEY (turma_id) REFERENCES turmas(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS escalas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            turma_id INTEGER NOT NULL,
            semana_numero INTEGER NOT NULL,
            data_inicio DATE NOT NULL,
            data_fim DATE NOT NULL,
            dupla TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (turma_id) REFERENCES turmas(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS escalas_historico (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            turma_id INTEGER NOT NULL,
            gerado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            dados TEXT NOT NULL,
            FOREIGN KEY (turma_id) REFERENCES turmas(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS avisos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT,
            body_markdown TEXT NOT NULL,
            body_html TEXT NOT NULL,
            start_at TIMESTAMP NOT NULL,
            end_at TIMESTAMP NOT NULL,
            is_shouting INTEGER DEFAULT 0,
            has_countdown INTEGER DEFAULT 0,
            active INTEGER DEFAULT 0,
            redirect_link TEXT,
            image_path TEXT,
            image_size TEXT DEFAULT 'medium',
            deleted_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alert_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alert_id INTEGER NOT NULL,
            event_type TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (alert_id) REFERENCES avisos(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS presenca (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data DATE NOT NULL,
            turma_id INTEGER NOT NULL,
            registros TEXT NOT NULL,
            responsavel TEXT,
            created_by TEXT,
            updated_by TEXT,
            created_lat REAL,
            created_lng REAL,
            created_loc_accuracy REAL,
            updated_lat REAL,
            updated_lng REAL,
            updated_loc_accuracy REAL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(data, turma_id),
            FOREIGN KEY (turma_id) REFERENCES turmas(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS historico_acoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario TEXT NOT NULL,
            acao TEXT NOT NULL,
            detalhes TEXT,
            turma TEXT,
            item_afetado TEXT,
            data_hora TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS auto_generation_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            turma_id INTEGER NOT NULL,
            data_geracao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            semanas_geradas INTEGER NOT NULL,
            status TEXT NOT NULL,
            FOREIGN KEY (turma_id) REFERENCES turmas(id)
        )
    ''')

    turmas_padrao = list(TURMAS.values())

    for turma in turmas_padrao:
        cursor.execute('INSERT OR IGNORE INTO turmas (nome) VALUES (?)', (turma,))

    conn.commit()
    conn.close()

    if not os.path.exists(UPLOAD_DIR):
        os.makedirs(UPLOAD_DIR)
    if not os.path.exists(THUMBNAIL_DIR):
        os.makedirs(THUMBNAIL_DIR)

# ============================================
# FUNÇÕES DE AUTENTICAÇÃO
# ============================================

def check_login_attempts(username, ip):
    key = f"{ip}:{username}"
    now = time.time()

    if key in login_attempts:
        attempts, first_attempt = login_attempts[key]
        if now - first_attempt > 900:
            login_attempts[key] = [1, now]
            return True

        if attempts >= 5:
            return False
        login_attempts[key] = [attempts + 1, first_attempt]
    else:
        login_attempts[key] = [1, now]

    return True

def record_failed_attempt(username, ip):
    key = f"{ip}:{username}"
    if key in login_attempts:
        attempts, first_attempt = login_attempts[key]
        login_attempts[key] = [attempts + 1, first_attempt]
    else:
        login_attempts[key] = [1, time.time()]

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        client_ip = get_remote_address()

        if not auth:
            return jsonify({'error': 'Acesso não autorizado'}), 401

        if not check_login_attempts(auth.username, client_ip):
            return jsonify({'error': 'Muitas tentativas de login. Aguarde 15 minutos.'}), 429

        admin = get_admin_by_username(auth.username)
        if admin:
            if check_password_hash(admin['password_hash'], auth.password):
                key = f"{client_ip}:{auth.username}"
                if key in login_attempts:
                    del login_attempts[key]
                return f(*args, **kwargs)

        record_failed_attempt(auth.username, client_ip)
        return jsonify({'error': 'Credenciais inválidas'}), 401

    return decorated

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Token não fornecido'}), 401

        token = auth_header.split(' ')[1]
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
            request.user = payload['username']
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Token expirado'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Token inválido'}), 401

        return f(*args, **kwargs)
    return decorated

def log_admin_action(usuario, acao, detalhes=None, turma=None, item_afetado=None):
    try:
        conn = get_db()
        cursor = conn.cursor()

        usuario = sanitizar_texto(usuario)
        acao = sanitizar_texto(acao)
        detalhes = sanitizar_texto(detalhes) if detalhes else None
        turma = sanitizar_texto(turma) if turma else None
        item_afetado = sanitizar_texto(item_afetado) if item_afetado else None

        cursor.execute('''
            INSERT INTO historico_acoes (usuario, acao, detalhes, turma, item_afetado)
            VALUES (?, ?, ?, ?, ?)
        ''', (usuario, acao, detalhes, turma, item_afetado))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Erro ao registrar ação: {e}")

# ============================================
# FUNÇÕES AUXILIARES
# ============================================

def get_turma_id_by_nome(turma_nome):
    turma_nome = sanitizar_texto(turma_nome)
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
    result = cursor.fetchone()
    conn.close()
    return result['id'] if result else None

def get_turma_nome_by_id(turma_id):
    try:
        turma_id = int(turma_id)
    except (ValueError, TypeError):
        return None

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT nome FROM turmas WHERE id = ?', (turma_id,))
    result = cursor.fetchone()
    conn.close()
    return result['nome'] if result else None

def get_all_turmas():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, nome FROM turmas ORDER BY nome')
    turmas = [{'id': row['id'], 'nome': row['nome']} for row in cursor.fetchall()]
    conn.close()
    return turmas

def formatar_data_brasil(data_obj):
    return data_obj.strftime('%d/%m')

def formatar_data_range_brasil(data_inicio, data_fim):
    return f"{data_inicio.strftime('%d/%m')} - {data_fim.strftime('%d/%m')}"

def obter_ultima_data_escala(turma_id):
    try:
        turma_id = int(turma_id)
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT MAX(data_fim) as ultima_data_fim
            FROM escalas
            WHERE turma_id = ?
        ''', (turma_id,))
        result = cursor.fetchone()
        conn.close()

        if result and result['ultima_data_fim']:
            return datetime.strptime(result['ultima_data_fim'], '%Y-%m-%d')
        return None
    except Exception as e:
        print(f"Erro ao obter última data da escala: {e}")
        return None

def obter_proximo_dia_util(data):
    data_atual = data
    while data_atual.weekday() >= 5:
        data_atual += timedelta(days=1)
    return data_atual

def get_semana_atual():
    hoje = datetime.now().date()

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT data_inicio FROM escalas
        WHERE turma_id = (SELECT id FROM turmas WHERE nome = ?)
        ORDER BY semana_numero LIMIT 1
    ''', (TURMAS['formare'],))
    result = cursor.fetchone()
    conn.close()

    if result and result['data_inicio']:
        data_inicio = datetime.strptime(result['data_inicio'], '%Y-%m-%d').date()
    else:
        data_inicio = datetime(2026, 4, 6).date()

    diff = (hoje - data_inicio).days
    semana = (diff // 7) + 1

    if semana < 1:
        semana = 1

    return f"Semana {semana:02d}: {hoje.strftime('%d/%m')}"

# ============================================
# FUNÇÃO DE GERAÇÃO DE ESCALAS
# ============================================

def gerar_escala_fair(turma_nome, force_new=False):
    turma_nome = sanitizar_texto(turma_nome)
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
    turma = cursor.fetchone()
    if not turma:
        conn.close()
        return None

    turma_id = turma['id']

    cursor.execute('''
        SELECT nome FROM alunos
        WHERE turma_id = ? AND is_admin = 0
    ''', (turma_id,))
    alunos = [row['nome'] for row in cursor.fetchall()]
    conn.close()

    if len(alunos) != 18:
        return None

    random.shuffle(alunos)

    semanas = []
    data_inicio = DATA_INICIO_ESCALA

    pares = []
    for i in range(0, len(alunos), 2):
        pares.append([alunos[i], alunos[i + 1]])

    random.shuffle(pares)

    for semana_num, dupla in enumerate(pares, 1):
        data_semana_inicio = data_inicio + timedelta(weeks=semana_num - 1)
        data_semana_fim = data_semana_inicio + timedelta(days=4)

        semanas.append({
            'semana_numero': semana_num,
            'data_inicio': data_semana_inicio.strftime('%Y-%m-%d'),
            'data_fim': data_semana_fim.strftime('%Y-%m-%d'),
            'dupla': dupla
        })

    return semanas

def salvar_escala(turma_nome, semanas):
    turma_nome = sanitizar_texto(turma_nome)
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
    turma = cursor.fetchone()
    if not turma:
        conn.close()
        return False

    turma_id = turma['id']

    cursor.execute('DELETE FROM escalas WHERE turma_id = ?', (turma_id,))

    for semana in semanas:
        cursor.execute('''
            INSERT INTO escalas (turma_id, semana_numero, data_inicio, data_fim, dupla)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            turma_id,
            semana['semana_numero'],
            semana['data_inicio'],
            semana['data_fim'],
            json.dumps(semana['dupla'])
        ))

    cursor.execute('''
        INSERT INTO escalas_historico (turma_id, dados)
        VALUES (?, ?)
    ''', (turma_id, json.dumps(semanas)))

    conn.commit()
    conn.close()
    return True

def verificar_e_gerar_nova_escala():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id, nome FROM turmas')
        turmas = cursor.fetchall()
        conn.close()

        resultados = []
        data_atual = datetime.now().date()

        if data_atual > DATA_LIMITE_AUTOMATICA.date():
            return False, "Período de geração automática encerrado (após 31/12/2026)"

        for turma in turmas:
            turma_id = turma['id']
            turma_nome = turma['nome']

            ultima_data_fim = obter_ultima_data_escala(turma_id)

            if not ultima_data_fim:
                continue

            if data_atual > ultima_data_fim.date():
                nova_data_inicio = obter_proximo_dia_util(ultima_data_fim + timedelta(days=1))

                if nova_data_inicio.date() > DATA_LIMITE_AUTOMATICA.date():
                    resultados.append(f"Turma {turma_nome}: limite de data atingido")
                    continue

                global DATA_INICIO_ESCALA
                data_inicio_original = DATA_INICIO_ESCALA
                DATA_INICIO_ESCALA = nova_data_inicio

                semanas = gerar_escala_fair(turma_nome, force_new=True)

                if semanas:
                    salvar_escala(turma_nome, semanas)

                    conn = get_db()
                    cursor = conn.cursor()
                    cursor.execute('''
                        INSERT INTO auto_generation_log (turma_id, semanas_geradas, status)
                        VALUES (?, ?, ?)
                    ''', (turma_id, len(semanas), 'success'))
                    conn.commit()
                    conn.close()

                    log_admin_action('Sistema', 'Gerar Escala Automática',
                                   f'Nova escala gerada automaticamente para {turma_nome} com {len(semanas)} semanas iniciando em {nova_data_inicio.strftime("%d/%m/%Y")}')

                    resultados.append(f"Turma {turma_nome}: nova escala gerada com {len(semanas)} semanas")
                else:
                    resultados.append(f"Turma {turma_nome}: falha ao gerar escala (verifique se há 18 alunos)")

                DATA_INICIO_ESCALA = data_inicio_original

        if resultados:
            return True, " | ".join(resultados)
        return False, "Nenhuma nova escala necessária no momento"

    except Exception as e:
        print(f"Erro na verificação automática de escalas: {e}")
        return False, f"Erro: {str(e)}"

# ============================================
# ROTAS DE LOGIN
# ============================================

@app.route('/api/admin/token', methods=['POST'])
def admin_token():
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        client_ip = get_remote_address()

        if not username or not password:
            return jsonify({'success': False, 'error': 'Usuário e senha são obrigatórios'}), 400

        if is_ip_blocked(client_ip):
            return jsonify({'success': False, 'error': 'IP bloqueado. Contate o administrador.'}), 403

        admin = get_admin_by_username(username)

        if admin:
            if check_password_hash(admin['password_hash'], password):
                # 🔒 BLOQUEAR SUPER ADMIN DE LOGAR NO PAINEL NORMAL
                if is_super_admin(username):
                    return jsonify({'success': False, 'error': 'Acesso negado. Este usuário só pode acessar o Super Admin.'}), 403

                # Resto do código continua igual...
                payload = {
                    'username': username,
                    'exp': datetime.utcnow() + timedelta(hours=JWT_EXPIRATION),
                    'iat': datetime.utcnow()
                }
                token = jwt.encode(payload, JWT_SECRET, algorithm='HS256')

                admin_type = admin['admin_type']
                turmas_permitidas = get_admin_turmas_permitidas(username)
                permissoes = get_admin_permissoes(username)

                log_admin_action(username, 'Login', f'Login via token - tipo: {admin_type}')
                log_admin_action_db(username, 'login', f'Login realizado - IP: {client_ip}', client_ip, request.headers.get('User-Agent'))

                return jsonify({
                    'success': True,
                    'token': token,
                    'admin': username,
                    'tipo_admin': admin_type,
                    'turmas_permitidas': turmas_permitidas,
                    'permissoes': permissoes,
                    'is_global_admin': admin_type == 'global',
                    'expires_in': JWT_EXPIRATION * 3600
                })

        return jsonify({'success': False, 'error': 'Credenciais inválidas'}), 401

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        client_ip = get_remote_address()

        if not username or not password:
            return jsonify({'success': False, 'error': 'Usuário e senha são obrigatórios'}), 400

        if is_ip_blocked(client_ip):
            return jsonify({'success': False, 'error': 'IP bloqueado. Contate o administrador.'}), 403

        admin = get_admin_by_username(username)

        if admin:
            if check_password_hash(admin['password_hash'], password):
                session['admin'] = username
                session.permanent = True

                admin_type = admin['admin_type']
                turmas_permitidas = get_admin_turmas_permitidas(username)
                permissoes = get_admin_permissoes(username)

                log_admin_action(username, 'Login (Session)', f'Login via sessão - tipo: {admin_type}')
                log_admin_action_db(username, 'login_session', f'Login via sessão - IP: {client_ip}', client_ip, request.headers.get('User-Agent'))

                return jsonify({
                    'success': True,
                    'admin': username,
                    'tipo_admin': admin_type,
                    'turmas_permitidas': turmas_permitidas,
                    'permissoes': permissoes,
                    'is_global_admin': admin_type == 'global',
                    'message': 'Login realizado com sucesso'
                })

        return jsonify({'success': False, 'error': 'Credenciais inválidas'}), 401

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/login', methods=['POST'])
@limiter.limit("10000 per minute")
def login_aluno():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Dados inválidos'}), 400

        turma_nome = data.get('turma')
        edv = data.get('edv')

        if not turma_nome or not edv:
            return jsonify({'success': False, 'error': 'Turma e EDV são obrigatórios'}), 400

        turma_nome = sanitizar_texto(turma_nome)
        edv = sanitizar_texto(edv)

        edv_permitido = None
        for permitido in EDVS_PERMITIDOS:
            if permitido["edv"] == edv:
                edv_permitido = permitido
                break

        if edv_permitido:
            session['user'] = {
                'nome': edv_permitido["nome"],
                'is_admin': True,
                'tipo_admin': 'formare',
                'turmas_permitidas': list(TURMAS.values()),
                'permissoes': PERMISSOES_ADMIN_FORMARE,
                'turma': turma_nome
            }
            session.permanent = True

            log_admin_action(edv_permitido["nome"], 'Login (Global Read-Only)',
                           f'Login global read-only realizado na turma {turma_nome}',
                           turma_nome, edv)

            return jsonify({
                'success': True,
                'nome': edv_permitido["nome"],
                'is_admin': False,
                'tipo_admin': 'formare',
                'turmas_permitidas': list(TURMAS.values()),
                'permissoes': PERMISSOES_ADMIN_FORMARE
            })

        if not validar_edv(edv):
            return jsonify({'success': False, 'error': 'EDV inválido'}), 400

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()
        if not turma:
            conn.close()
            return jsonify({'success': False, 'error': 'Turma não encontrada'}), 404

        cursor.execute('''
            SELECT nome, is_admin FROM alunos
            WHERE turma_id = ? AND edv = ?
        ''', (turma['id'], edv))

        aluno = cursor.fetchone()

        if not aluno:
            cursor.execute('''
                SELECT nome, is_admin FROM alunos
                WHERE edv = ? AND is_admin = 1
            ''', (edv,))
            aluno = cursor.fetchone()

            if aluno:
                turma_nome = data.get('turma')

                admin_type = get_admin_type(aluno['nome'])
                turmas_permitidas = get_admin_turmas_permitidas(aluno['nome'])
                permissoes = get_admin_permissoes(aluno['nome'])

                session['user'] = {
                    'nome': aluno['nome'],
                    'is_admin': True,
                    'tipo_admin': admin_type,
                    'turmas_permitidas': turmas_permitidas,
                    'permissoes': permissoes,
                    'turma': turma_nome
                }
                session.permanent = True

                log_admin_action(aluno['nome'], 'Login (Admin)',
                               f'Login de administrador tipo {admin_type} na turma {turma_nome}',
                               turma_nome, edv)

                return jsonify({
                    'success': True,
                    'nome': aluno['nome'],
                    'is_admin': True,
                    'tipo_admin': admin_type,
                    'turmas_permitidas': turmas_permitidas,
                    'permissoes': permissoes
                })

        conn.close()

        if aluno:
            session['user'] = {
                'nome': aluno['nome'],
                'is_admin': False,
                'tipo_admin': None,
                'turmas_permitidas': [turma_nome],
                'permissoes': [],
                'turma': turma_nome
            }
            session.permanent = True

            log_admin_action(aluno['nome'], 'Login',
                           f'Login realizado na turma {turma_nome}',
                           turma_nome, edv)

            return jsonify({
                'success': True,
                'nome': aluno['nome'],
                'is_admin': False,
                'tipo_admin': None,
                'turmas_permitidas': [turma_nome],
                'permissoes': []
            })
        else:
            return jsonify({'success': False, 'error': 'EDV não encontrado para esta turma'}), 404

    except Exception as e:
        return jsonify({'success': False, 'error': 'Erro interno'}), 500

@app.route('/api/me', methods=['GET'])
def get_current_user():
    try:
        if 'admin' in session:
            admin_type = get_admin_type(session['admin'])
            turmas_permitidas = get_admin_turmas_permitidas(session['admin'])
            permissoes = get_admin_permissoes(session['admin'])

            return jsonify({
                'success': True,
                'nome': session['admin'],
                'is_admin': True,
                'tipo_admin': admin_type,
                'turmas_permitidas': turmas_permitidas,
                'permissoes': permissoes
            })

        if 'user' in session:
            return jsonify({
                'success': True,
                'nome': session['user']['nome'],
                'is_admin': session['user']['is_admin'],
                'tipo_admin': session['user'].get('tipo_admin'),
                'turmas_permitidas': session['user'].get('turmas_permitidas', []),
                'permissoes': session['user'].get('permissoes', [])
            })

        return jsonify({'success': False}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    try:
        session.clear()
        return jsonify({'success': True, 'message': 'Logout realizado com sucesso'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================
# ROTAS DE PERMISSÕES DO USUÁRIO
# ============================================

@app.route('/api/admin/permissions', methods=['GET'])
@token_required
def get_user_permissions():
    """Retorna as permissões detalhadas do usuário atual"""
    try:
        username = request.user
        admin_type = get_admin_type(username)
        turmas_permitidas = get_admin_turmas_permitidas(username)
        permissoes = get_admin_permissoes(username)

        pode_fazer_presenca = can_perform_action(username, 'presenca', None)
        pode_fazer_upload = can_perform_action(username, 'upload', None)
        pode_fazer_avisos = can_perform_action(username, 'avisos', None)
        pode_fazer_backup = can_perform_action(username, 'backup', None)
        pode_gerenciar_alunos = can_perform_action(username, 'alunos', None)
        pode_ver_historico = can_perform_action(username, 'historico', None)

        return jsonify({
            'success': True,
            'username': username,
            'tipo_admin': admin_type,
            'turmas_permitidas': turmas_permitidas,
            'permissoes': permissoes,
            'detalhes_permissoes': {
                'pode_fazer_presenca': pode_fazer_presenca,
                'pode_fazer_upload': pode_fazer_upload,
                'pode_fazer_avisos': pode_fazer_avisos,
                'pode_fazer_backup': pode_fazer_backup,
                'pode_gerenciar_alunos': pode_gerenciar_alunos,
                'pode_ver_historico': pode_ver_historico
            }
        })

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ROTA PARA LISTAR TODAS AS TURMAS
# ============================================

@app.route('/api/turmas', methods=['GET'])
@token_required
def list_all_turmas():
    try:
        username = request.user
        turmas_permitidas = get_admin_turmas_permitidas(username)

        if get_admin_type(username) in ['global', 'formare']:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT id, nome FROM turmas ORDER BY nome')
            turmas = [{'id': row['id'], 'nome': row['nome']} for row in cursor.fetchall()]
            conn.close()
        else:
            turmas = [{'id': get_turma_id_by_nome(t), 'nome': t} for t in turmas_permitidas if get_turma_id_by_nome(t)]

        return jsonify({
            'success': True,
            'turmas': turmas
        })
    except Exception as e:
        print(f"Erro ao listar turmas: {e}")
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ROTAS DE ESCALAS
# ============================================

@app.route('/api/escala/<path:turma_nome>', methods=['GET'])
def get_escala(turma_nome):
    """Rota PÚBLICA para alunos visualizarem a escala"""
    try:
        turma_nome = sanitizar_texto(turma_nome)

        is_authorized = False

        if 'admin' in session:
            username = session['admin']
            is_authorized = can_visualize_turma(username, turma_nome)
        elif 'user' in session:
            username = session['user'].get('nome')
            user_turma = session['user'].get('turma')
            if user_turma == turma_nome:
                is_authorized = True

        if not is_authorized:
            conn_check = get_db()
            cursor_check = conn_check.cursor()
            cursor_check.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
            turma_exists = cursor_check.fetchone()
            conn_check.close()

            if not turma_exists:
                return jsonify({'error': 'Turma não encontrada'}), 404
            is_authorized = True

        verificar_e_gerar_nova_escala()

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()
        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('''
            SELECT semana_numero, data_inicio, data_fim, dupla
            FROM escalas
            WHERE turma_id = ?
            ORDER BY semana_numero
        ''', (turma['id'],))

        escalas = cursor.fetchall()
        conn.close()

        if not escalas:
            return jsonify({'semana_atual': get_semana_atual(), 'duplas': []})

        duplas = []
        for escala in escalas:
            dupla_nomes = json.loads(escala['dupla'])
            data_inicio_obj = datetime.strptime(escala['data_inicio'], '%Y-%m-%d')
            data_fim_obj = datetime.strptime(escala['data_fim'], '%Y-%m-%d')
            data_range = formatar_data_range_brasil(data_inicio_obj, data_fim_obj)

            duplas.append({
                'semana': f"Semana {escala['semana_numero']:02d}: {data_range}",
                'dupla': f"{dupla_nomes[0]} e {dupla_nomes[1]}"
            })

        return jsonify({
            'semana_atual': get_semana_atual(),
            'duplas': duplas
        })

    except Exception as e:
        print(f"Erro em get_escala: {e}")
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/escalas/<path:turma_nome>', methods=['GET'])
@token_required
def get_escala_admin(turma_nome):
    try:
        if not can_visualize_turma(request.user, turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        turma_nome = sanitizar_texto(turma_nome)
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()
        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('''
            SELECT semana_numero, data_inicio, data_fim, dupla
            FROM escalas
            WHERE turma_id = ?
            ORDER BY semana_numero
        ''', (turma['id'],))

        escalas = cursor.fetchall()
        conn.close()

        duplas = []
        for escala in escalas:
            dupla = json.loads(escala['dupla'])
            data_inicio_obj = datetime.strptime(escala['data_inicio'], '%Y-%m-%d')
            data_fim_obj = datetime.strptime(escala['data_fim'], '%Y-%m-%d')
            data_range = formatar_data_range_brasil(data_inicio_obj, data_fim_obj)

            duplas.append({
                'semana': f"Semana {escala['semana_numero']:02d}: {data_range}",
                'dupla': f"{dupla[0]} e {dupla[1]}"
            })

        return jsonify({
            'semana_atual': get_semana_atual(),
            'duplas': duplas
        })

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/escalas/<path:turma_nome>', methods=['PUT'])
@token_required
def update_escala(turma_nome):
    try:
        print(f"🔵 [DEBUG] Iniciando update_escala para turma: {turma_nome}")
        print(f"🔵 [DEBUG] Usuário: {request.user}")

        if not can_perform_action(request.user, 'escalas', turma_nome):
            print(f"🔴 [DEBUG] Acesso negado para {request.user}")
            return jsonify({'error': 'Acesso negado'}), 403

        data = request.get_json()
        print(f"🔵 [DEBUG] Dados recebidos: {data}")

        duplas = data.get('duplas')
        print(f"🔵 [DEBUG] Duplas: {duplas}")

        if not duplas:
            print(f"🔴 [DEBUG] Nenhuma dupla recebida")
            return jsonify({'error': 'Dados da escala são obrigatórios'}), 400

        turma_nome = sanitizar_texto(turma_nome)
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()
        if not turma:
            conn.close()
            print(f"🔴 [DEBUG] Turma não encontrada: {turma_nome}")
            return jsonify({'error': 'Turma não encontrada'}), 404

        print(f"🔵 [DEBUG] Turma ID: {turma['id']}")

        # Deletar escalas existentes
        cursor.execute('DELETE FROM escalas WHERE turma_id = ?', (turma['id'],))
        print(f"🔵 [DEBUG] Escalas antigas deletadas")

        for semana_num, dupla_info in enumerate(duplas, 1):
            dupla_texto = dupla_info.get('dupla', '')
            print(f"🔵 [DEBUG] Processando semana {semana_num}: '{dupla_texto}'")

            # CORREÇÃO: Divide apenas no último " e "
            if ' e ' in dupla_texto:
                ultimo_e = dupla_texto.rfind(' e ')
                nome1 = dupla_texto[:ultimo_e].strip()
                nome2 = dupla_texto[ultimo_e + 3:].strip()
                dupla_nomes = [nome1, nome2]
                print(f"🔵 [DEBUG] Nomes separados: '{nome1}' e '{nome2}'")
            else:
                print(f"🔴 [DEBUG] Formato inválido, sem ' e ': '{dupla_texto}'")
                continue

            if len(dupla_nomes) != 2:
                print(f"🔴 [DEBUG] Não tem 2 nomes: {dupla_nomes}")
                continue

            # Usa a data digitada no front
            semana_texto = dupla_info.get('semana', '')
            print(f"🔵 [DEBUG] Semana recebida: {semana_texto}")

            try:
                # Ex.: "Semana 01: 22/05 - 22/05"
                if ':' in semana_texto:
                    semana_texto = semana_texto.split(':', 1)[1].strip()

                inicio_str, fim_str = [
                    p.strip()
                    for p in semana_texto.split('-')
                ]

                ano = datetime.now().year

                data_inicio = datetime.strptime(
                    f"{inicio_str}/{ano}",
                    "%d/%m/%Y"
                )

                data_fim = datetime.strptime(
                    f"{fim_str}/{ano}",
                    "%d/%m/%Y"
                )

            except Exception as e:
                print(f"🔴 [DEBUG] Erro ao interpretar data: {e}")

                # fallback
                data_inicio = DATA_INICIO_ESCALA + timedelta(
                    weeks=semana_num - 1
                )
                data_fim = data_inicio

            print(f"🔵 [DEBUG] Inserindo: semana={semana_num}, inicio={data_inicio}, fim={data_fim}, dupla={dupla_nomes}")

            cursor.execute('''
                INSERT INTO escalas (turma_id, semana_numero, data_inicio, data_fim, dupla)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                turma['id'],
                semana_num,
                data_inicio.strftime('%Y-%m-%d'),
                data_fim.strftime('%Y-%m-%d'),
                json.dumps(dupla_nomes)
            ))

        conn.commit()

        # Verificar se salvou
        cursor.execute('SELECT COUNT(*) as total FROM escalas WHERE turma_id = ?', (turma['id'],))
        count = cursor.fetchone()
        print(f"🔵 [DEBUG] Total de escalas salvas: {count['total'] if count else 0}")

        conn.close()

        log_admin_action(request.user, 'Atualizar Escala', f'Escala da turma {turma_nome} atualizada', turma_nome)

        return jsonify({'success': True, 'message': f'Escala salva com sucesso. {count["total"] if count else 0} semanas salvas.'})

    except Exception as e:
        print(f"🔴 [DEBUG] ERRO: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
@app.route('/api/admin/escalas/<path:turma_nome>/generate', methods=['POST'])
@token_required
def generate_escala(turma_nome):
    try:
        if not can_perform_action(request.user, 'escalas', turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        turma_nome = sanitizar_texto(turma_nome)
        semanas = gerar_escala_fair(turma_nome)

        if not semanas:
            return jsonify({'error': 'Não foi possível gerar a escala. Verifique se há exatamente 18 alunos cadastrados (não admins)'}), 400

        salvar_escala(turma_nome, semanas)

        log_admin_action(request.user, 'Gerar Escala', f'Nova escala gerada para {turma_nome}', turma_nome)

        return jsonify({'success': True, 'message': f'Escala gerada com {len(semanas)} semanas'})

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/auto-generate', methods=['POST'])
@token_required
def trigger_auto_generate():
    try:
        if get_admin_type(request.user) != 'global':
            return jsonify({'error': 'Acesso negado. Apenas administradores globais podem disparar geração automática.'}), 403

        gerou, mensagem = verificar_e_gerar_nova_escala()
        return jsonify({
            'success': gerou,
            'message': mensagem
        })
    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/scale-status', methods=['GET'])
@token_required
def get_scale_status():
    try:
        if get_admin_type(request.user) not in ['global', 'formare']:
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id, nome FROM turmas')
        turmas = cursor.fetchall()
        conn.close()

        data_atual = datetime.now().date()
        resultados = []

        for turma in turmas:
            turma_id = turma['id']
            turma_nome = turma['nome']

            ultima_data_fim = obter_ultima_data_escala(turma_id)

            if ultima_data_fim:
                if data_atual > ultima_data_fim.date():
                    status = "expirada"
                    dias_atraso = (data_atual - ultima_data_fim.date()).days
                    detalhe = f"Atrasada há {dias_atraso} dias"
                else:
                    status = "ativa"
                    dias_restantes = (ultima_data_fim.date() - data_atual).days
                    detalhe = f"Termina em {dias_restantes} dias"
            else:
                status = "sem_escala"
                detalhe = "Nenhuma escala cadastrada"

            resultados.append({
                'turma': turma_nome,
                'status': status,
                'detalhe': detalhe,
                'ultima_data_fim': ultima_data_fim.strftime('%d/%m/%Y') if ultima_data_fim else None
            })

        return jsonify({
            'data_atual': data_atual.strftime('%d/%m/%Y'),
            'limite_automatico': DATA_LIMITE_AUTOMATICA.strftime('%d/%m/%Y'),
            'escalas': resultados
        })

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ROTAS DE ALUNOS (ADMIN)
# ============================================

@app.route('/api/admin/turmas/<path:turma_nome>/alunos', methods=['GET'])
@token_required
def list_alunos(turma_nome):
    try:
        turma_nome = sanitizar_texto(turma_nome)

        if not can_visualize_turma(request.user, turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()
        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('''
            SELECT edv, nome FROM alunos
            WHERE turma_id = ? AND is_admin = 0
            ORDER BY nome
        ''', (turma['id'],))

        alunos = {}
        for row in cursor.fetchall():
            alunos[row['edv']] = row['nome']

        conn.close()
        return jsonify(alunos)

    except Exception as e:
        print(f"Erro ao listar alunos: {e}")
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/turmas/<path:turma_nome>/alunos/all', methods=['GET'])
@token_required
def list_all_users(turma_nome):
    try:
        turma_nome = sanitizar_texto(turma_nome)

        admin_type = get_admin_type(request.user)
        if admin_type not in ['global', 'formare']:
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()
        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('''
            SELECT edv, nome, is_admin FROM alunos
            WHERE turma_id = ?
            ORDER BY nome
        ''', (turma['id'],))

        alunos = {}
        for row in cursor.fetchall():
            nome = row['nome']
            if row['is_admin']:
                if '👩‍🏫' not in nome and '👨‍🏫' not in nome:
                    nome = nome + ('👩‍🏫' if 'a' in nome[-2:] else '👨‍🏫')
            alunos[row['edv']] = nome

        conn.close()
        return jsonify(alunos)

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/turmas/<path:turma_nome>/alunos', methods=['POST'])
@token_required
def add_aluno(turma_nome):
    try:
        if not can_perform_action(request.user, 'alunos', turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        data = request.get_json()
        edv = data.get('edv')
        nome = data.get('nome')
        is_admin = data.get('is_admin', 0)

        if not edv or not nome:
            return jsonify({'error': 'EDV e nome são obrigatórios'}), 400

        edv = sanitizar_texto(edv)
        nome = sanitizar_texto(nome)
        turma_nome = sanitizar_texto(turma_nome)

        if not validar_edv(edv):
            return jsonify({'error': 'EDV deve conter apenas números'}), 400

        if not validar_nome(nome):
            return jsonify({'error': 'Nome inválido'}), 400

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()
        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('SELECT id FROM alunos WHERE turma_id = ? AND edv = ?', (turma['id'], edv))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'EDV já cadastrado nesta turma'}), 409

        cursor.execute('''
            INSERT INTO alunos (edv, nome, turma_id, is_admin)
            VALUES (?, ?, ?, ?)
        ''', (edv, nome, turma['id'], 1 if is_admin else 0))

        conn.commit()
        conn.close()

        log_admin_action(request.user, 'Adicionar Aluno', f'Adicionado {"administrador" if is_admin else "aluno"} {nome} (EDV: {edv})', turma_nome, edv)

        return jsonify({'success': True, 'message': f'{"Administrador" if is_admin else "Aluno"} adicionado com sucesso'})

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/turmas/<path:turma_nome>/alunos', methods=['DELETE'])
@token_required
def remove_aluno(turma_nome):
    try:
        if not can_perform_action(request.user, 'alunos', turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        edv = request.args.get('edv')

        if not edv:
            return jsonify({'error': 'EDV é obrigatório'}), 400

        edv = sanitizar_texto(edv)
        turma_nome = sanitizar_texto(turma_nome)

        if not validar_edv(edv):
            return jsonify({'error': 'EDV inválido'}), 400

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()
        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('SELECT nome, is_admin FROM alunos WHERE turma_id = ? AND edv = ?', (turma['id'], edv))
        aluno = cursor.fetchone()

        if not aluno:
            conn.close()
            return jsonify({'error': 'Aluno não encontrado'}), 404

        cursor.execute('DELETE FROM alunos WHERE turma_id = ? AND edv = ?', (turma['id'], edv))
        conn.commit()
        conn.close()

        log_admin_action(request.user, 'Remover Aluno', f'Removido {"administrador" if aluno["is_admin"] else "aluno"} {aluno["nome"]} (EDV: {edv})', turma_nome, edv)

        return jsonify({'success': True, 'message': 'Aluno removido com sucesso'})

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/alunos/import', methods=['POST'])
@token_required
def import_alunos():
    try:
        turma_nome = request.form.get('turma', TURMAS['formare'])
        if not can_perform_action(request.user, 'alunos', turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400

        file = request.files['file']
        turma_nome = sanitizar_texto(turma_nome)

        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo vazio'}), 400

        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if ext not in ['csv', 'xlsx', 'xls']:
            return jsonify({'error': 'Tipo de arquivo não permitido. Use CSV ou Excel'}), 400

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()
        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        turma_id = turma['id']

        content = file.read()
        alunos_importados = []
        erros = []

        try:
            text_content = content.decode('utf-8')
            csv_reader = csv.reader(io.StringIO(text_content))
            next(csv_reader, None)

            for row_num, row in enumerate(csv_reader, start=2):
                if len(row) < 2:
                    erros.append(f"Linha {row_num}: formato inválido")
                    continue

                edv = row[0].strip()
                nome = row[1].strip()

                if not edv or not nome:
                    erros.append(f"Linha {row_num}: EDV ou nome vazio")
                    continue

                if not validar_edv(edv):
                    erros.append(f"Linha {row_num}: EDV inválido (apenas números)")
                    continue

                if not validar_nome(nome):
                    erros.append(f"Linha {row_num}: Nome inválido")
                    continue

                try:
                    cursor.execute('SELECT id FROM alunos WHERE turma_id = ? AND edv = ?', (turma_id, edv))
                    if cursor.fetchone():
                        erros.append(f"Linha {row_num}: EDV {edv} já cadastrado")
                        continue

                    cursor.execute('''
                        INSERT INTO alunos (edv, nome, turma_id, is_admin)
                        VALUES (?, ?, ?, 0)
                    ''', (edv, nome, turma_id))
                    alunos_importados.append(nome)

                except Exception as e:
                    erros.append(f"Linha {row_num}: {str(e)}")

        except UnicodeDecodeError:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or len(row) < 2:
                        continue

                    edv = str(row[0]).strip() if row[0] else ''
                    nome = str(row[1]).strip() if row[1] else ''

                    if not edv or not nome:
                        erros.append(f"Linha {row_num}: EDV ou nome vazio")
                        continue

                    if not validar_edv(edv):
                        erros.append(f"Linha {row_num}: EDV inválido (apenas números)")
                        continue

                    if not validar_nome(nome):
                        erros.append(f"Linha {row_num}: Nome inválido")
                        continue

                    try:
                        cursor.execute('SELECT id FROM alunos WHERE turma_id = ? AND edv = ?', (turma_id, edv))
                        if cursor.fetchone():
                            erros.append(f"Linha {row_num}: EDV {edv} já cadastrado")
                            continue

                        cursor.execute('''
                            INSERT INTO alunos (edv, nome, turma_id, is_admin)
                            VALUES (?, ?, ?, 0)
                        ''', (edv, nome, turma_id))
                        alunos_importados.append(nome)

                    except Exception as e:
                        erros.append(f"Linha {row_num}: {str(e)}")

            except Exception as e:
                return jsonify({'error': f'Erro ao processar arquivo: formato inválido'}), 400

        conn.commit()
        conn.close()
        log_admin_action(request.user, 'Importar Alunos',
                        f'Importados {len(alunos_importados)} alunos na turma {turma_nome}. Erros: {len(erros)}',
                        turma_nome)

        return jsonify({
            'success': True,
            'importados': len(alunos_importados),
            'erros': erros[:10],
            'message': f'{len(alunos_importados)} alunos importados com sucesso'
        })

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ROTAS DE PRESENÇA
# ============================================

@app.route('/api/presenca/salvar', methods=['POST'])
@token_required
def save_attendance():
    try:
        data = request.get_json()
        data_str = data.get('data')
        registros = data.get('registros')
        turma_nome = data.get('turma', TURMAS['formare'])

        responsavel = request.user if hasattr(request, 'user') else 'Sistema'
        localizacao = data.get('localizacao')
        created_lat = created_lng = created_loc_accuracy = None
        updated_lat = updated_lng = updated_loc_accuracy = None

        if isinstance(localizacao, dict):
            try:
                lat = float(localizacao.get('lat'))
                lng = float(localizacao.get('lng'))
                accuracy = localizacao.get('accuracy')
                if -90 <= lat <= 90 and -180 <= lng <= 180:
                    created_lat = updated_lat = lat
                    created_lng = updated_lng = lng
                    created_loc_accuracy = updated_loc_accuracy = float(accuracy) if accuracy is not None else None
                else:
                    created_lat = created_lng = created_loc_accuracy = None
                    updated_lat = updated_lng = updated_loc_accuracy = None
            except (TypeError, ValueError):
                created_lat = created_lng = created_loc_accuracy = None
                updated_lat = updated_lng = updated_loc_accuracy = None

        if not data_str or not registros:
            return jsonify({'error': 'Data e registros são obrigatórios'}), 400

        turma_nome = sanitizar_texto(turma_nome)
        responsavel = sanitizar_texto(responsavel)

        if not can_perform_action(request.user, 'presenca', turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()

        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        for registro in registros:
            if 'status' not in registro or registro['status'] not in ['PRESENTE', 'AUSENTE', 'ATESTADO']:
                conn.close()
                return jsonify({'error': 'Status inválido'}), 400

        cursor.execute('SELECT id, created_by, updated_by FROM presenca WHERE data = ? AND turma_id = ?', (data_str, turma['id']))
        existing = cursor.fetchone()

        # Exigir coordenadas: se for atualização, updated_lat/updated_lng obrigatórios;
        # se for criação, created_lat/created_lng obrigatórios.
        if existing:
            if updated_lat is None or updated_lng is None:
                conn.close()
                return jsonify({'error': 'Localização (latitude/longitude) é obrigatória para atualizar a chamada.'}), 400
        else:
            if created_lat is None or created_lng is None:
                conn.close()
                return jsonify({'error': 'Localização (latitude/longitude) é obrigatória para salvar a chamada.'}), 400

        if existing:
            cursor.execute('''
                UPDATE presenca
                SET registros = ?, updated_at = CURRENT_TIMESTAMP, updated_by = ?,
                    updated_lat = ?, updated_lng = ?, updated_loc_accuracy = ?
                WHERE id = ?
            ''', (
                json.dumps(registros), responsavel,
                updated_lat, updated_lng, updated_loc_accuracy,
                existing['id']
            ))

            conn.commit()

            cursor.execute('SELECT created_by, updated_by, created_lat, created_lng, created_loc_accuracy, updated_lat, updated_lng, updated_loc_accuracy FROM presenca WHERE id = ?', (existing['id'],))
            result = cursor.fetchone()
            conn.close()

            return jsonify({
                'success': True,
                'message': 'Chamada atualizada com sucesso',
                'created_by': result['created_by'],
                'updated_by': result['updated_by'],
                'created_lat': result['created_lat'],
                'created_lng': result['created_lng'],
                'created_loc_accuracy': result['created_loc_accuracy'],
                'updated_lat': result['updated_lat'],
                'updated_lng': result['updated_lng'],
                'updated_loc_accuracy': result['updated_loc_accuracy'],
                'is_update': True
            })
        else:
            cursor.execute('''
                INSERT INTO presenca (data, turma_id, registros, responsavel, created_by, updated_by, created_lat, created_lng, created_loc_accuracy)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data_str, turma['id'], json.dumps(registros), responsavel,
                responsavel, None,
                created_lat, created_lng, created_loc_accuracy
            ))

            conn.commit()
            cursor.execute('SELECT created_by, updated_by, created_lat, created_lng, created_loc_accuracy, updated_lat, updated_lng, updated_loc_accuracy FROM presenca WHERE id = ?', (cursor.lastrowid,))
            result = cursor.fetchone()
            conn.close()

            return jsonify({
                'success': True,
                'message': 'Chamada salva com sucesso',
                'created_by': result['created_by'],
                'updated_by': result['updated_by'],
                'created_lat': result['created_lat'],
                'created_lng': result['created_lng'],
                'created_loc_accuracy': result['created_loc_accuracy'],
                'updated_lat': result['updated_lat'],
                'updated_lng': result['updated_lng'],
                'updated_loc_accuracy': result['updated_loc_accuracy'],
                'is_update': False
            })

    except Exception as e:
        print(f"Erro ao salvar chamada: {e}")
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/presenca/obter/<data_str>', methods=['GET'])
@token_required
def get_attendance(data_str):
    try:
        turma_nome = request.args.get('turma', TURMAS['formare'])
        turma_nome = sanitizar_texto(turma_nome)

        if not can_visualize_turma(request.user, turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()

        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('''
            SELECT registros, responsavel, created_at, updated_at, created_by, updated_by,
                   created_lat, created_lng, created_loc_accuracy,
                   updated_lat, updated_lng, updated_loc_accuracy
            FROM presenca
            WHERE data = ? AND turma_id = ?
        ''', (data_str, turma['id']))

        result = cursor.fetchone()
        conn.close()

        if result:
            registros = json.loads(result['registros'])
            return jsonify({
                'registros': registros,
                'created_by': result['created_by'],
                'updated_by': result['updated_by'],
                'created_at': result['created_at'],
                'updated_at': result['updated_at'],
                'created_lat': result['created_lat'],
                'created_lng': result['created_lng'],
                'created_loc_accuracy': result['created_loc_accuracy'],
                'updated_lat': result['updated_lat'],
                'updated_lng': result['updated_lng'],
                'updated_loc_accuracy': result['updated_loc_accuracy']
            })
        else:
            return jsonify({'registros': []}), 404

    except Exception as e:
        print(f"Erro ao obter chamada: {e}")
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/presenca/responsavel/<data_str>', methods=['GET'])
@token_required
def get_attendance_responsavel(data_str):
    try:
        turma_nome = request.args.get('turma', TURMAS['formare'])
        turma_nome = sanitizar_texto(turma_nome)

        if not can_visualize_turma(request.user, turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()

        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('''
            SELECT responsavel, created_at, updated_at
            FROM presenca
            WHERE data = ? AND turma_id = ?
        ''', (data_str, turma['id']))

        result = cursor.fetchone()
        conn.close()

        if result:
            return jsonify({
                'data': data_str,
                'responsavel': result['responsavel'],
                'criado_em': result['created_at'],
                'atualizado_em': result['updated_at']
            })
        else:
            return jsonify({'error': 'Nenhuma chamada encontrada para esta data'}), 404

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/presenca/responsaveis', methods=['GET'])
@token_required
def get_all_attendance_responsaveis():
    try:
        turma_nome = request.args.get('turma', TURMAS['formare'])
        turma_nome = sanitizar_texto(turma_nome)

        if not can_visualize_turma(request.user, turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()

        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('''
            SELECT data, responsavel, created_at, updated_at
            FROM presenca
            WHERE turma_id = ?
            ORDER BY data DESC
        ''', (turma['id'],))

        results = cursor.fetchall()
        conn.close()

        chamadas = []
        for row in results:
            chamadas.append({
                'data': row['data'],
                'responsavel': row['responsavel'],
                'criado_em': row['created_at'],
                'atualizado_em': row['updated_at']
            })

        return jsonify(chamadas)

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/presenca/meses_disponiveis', methods=['GET'])
def get_available_months():
    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT DISTINCT strftime('%Y-%m', data) as mes
            FROM presenca
            ORDER BY mes DESC
        ''')

        meses = [row['mes'] for row in cursor.fetchall()]
        conn.close()

        return jsonify({'meses': meses})

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ENDPOINT: ESTATÍSTICAS MENSAIS PARA GRÁFICO
# ============================================

@app.route('/api/presenca/estatisticas_mensais', methods=['GET'])
@token_required
def get_monthly_stats():
    try:
        mes = request.args.get('mes')
        turma_nome = request.args.get('turma', TURMAS['formare'])

        if not mes:
            return jsonify({'error': 'Mês é obrigatório (formato YYYY-MM)'}), 400

        turma_nome = sanitizar_texto(turma_nome)

        if not can_visualize_turma(request.user, turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        if not re.match(r'^\d{4}-\d{2}$', mes):
            return jsonify({'error': 'Formato de mês inválido. Use YYYY-MM'}), 400

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()

        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        turma_id = turma['id']

        cursor.execute('''
            SELECT data, registros FROM presenca
            WHERE turma_id = ? AND strftime('%Y-%m', data) = ?
        ''', (turma_id, mes))

        registros_mensais = cursor.fetchall()

        cursor.execute('''
            SELECT nome FROM alunos
            WHERE turma_id = ? AND is_admin = 0
            ORDER BY nome
        ''', (turma_id,))

        alunos = [row['nome'] for row in cursor.fetchall()]
        conn.close()

        if not alunos:
            return jsonify({'error': 'Nenhum aluno encontrado'}), 404

        presencas_por_data = {}

        for registro in registros_mensais:
            data = registro['data']
            registros = json.loads(registro['registros'])

            presencas_por_data[data] = {}
            for r in registros:
                nome_limpo = r['name'].replace('👩‍🏫', '').replace('👨‍🏫', '').strip()
                presencas_por_data[data][nome_limpo] = r['status']

        total_presentes = 0
        total_ausentes = 0
        total_atestados = 0
        total_registros = 0

        for data, presencas in presencas_por_data.items():
            for aluno in alunos:
                status = presencas.get(aluno)
                if status:
                    total_registros += 1
                    if status == 'PRESENTE':
                        total_presentes += 1
                    elif status == 'AUSENTE':
                        total_ausentes += 1
                    elif status == 'ATESTADO':
                        total_atestados += 1

        percentual_presente = 0
        percentual_ausente = 0
        percentual_atestado = 0
        percentual_nao_presente = 0

        if total_registros > 0:
            percentual_presente = round((total_presentes / total_registros) * 100, 1)
            percentual_ausente = round((total_ausentes / total_registros) * 100, 1)
            percentual_atestado = round((total_atestados / total_registros) * 100, 1)
            percentual_nao_presente = round(((total_ausentes + total_atestados) / total_registros) * 100, 1)

        return jsonify({
            'success': True,
            'mes': mes,
            'turma': turma_nome,
            'total_registros': total_registros,
            'presentes': total_presentes,
            'ausentes': total_ausentes,
            'atestados': total_atestados,
            'nao_presentes': total_ausentes + total_atestados,
            'percentual_presente': percentual_presente,
            'percentual_ausente': percentual_ausente,
            'percentual_atestado': percentual_atestado,
            'percentual_nao_presente': percentual_nao_presente,
            'dias_com_registro': len(presencas_por_data)
        })

    except Exception as e:
        print(f"Erro ao buscar estatísticas mensais: {e}")
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ROTAS DE AVISOS
# ============================================

@app.route('/api/public/alerts/active', methods=['GET'])
def get_active_alert():
    try:
        agora = datetime.now().isoformat()
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT * FROM avisos
            WHERE active = 1
            AND deleted_at IS NULL
            AND start_at <= ?
            AND end_at >= ?
            ORDER BY created_at DESC LIMIT 1
        ''', (agora, agora))

        aviso = cursor.fetchone()
        conn.close()

        if aviso:
            return jsonify({
                'id': aviso['id'],
                'title': sanitizar_texto(aviso['title']) if aviso['title'] else None,
                'body_html': sanitizar_html(aviso['body_html']),
                'image_path': aviso['image_path'],
                'has_countdown': bool(aviso['has_countdown']),
                'countdown_ends_at': aviso['end_at'],
                'redirect_link': aviso['redirect_link'],
                'is_shouting': bool(aviso['is_shouting']),
                'active': True
            })
        else:
            return jsonify({'active': False})

    except Exception as e:
        print(f"Erro ao buscar aviso ativo: {e}")
        return jsonify({'active': False}), 200

@app.route('/api/public/alerts/<int:alert_id>/event', methods=['POST'])
def register_alert_event(alert_id):
    try:
        data = request.get_json() or {}
        event_type = data.get('event_type')
        client_ip = get_remote_address()

        if event_type not in ['view', 'click', 'close']:
            return jsonify({'error': 'Tipo de evento inválido'}), 400

        if event_type == 'view':
            if not hasattr(app, '_viewed_alerts'):
                app._viewed_alerts = {}

            if alert_id not in app._viewed_alerts:
                app._viewed_alerts[alert_id] = set()

            if client_ip in app._viewed_alerts[alert_id]:
                return jsonify({'success': True, 'duplicate': True})

            app._viewed_alerts[alert_id].add(client_ip)

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO alert_events (alert_id, event_type)
            VALUES (?, ?)
        ''', (alert_id, event_type))
        conn.commit()
        conn.close()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/alerts', methods=['GET'])
@token_required
def get_alerts():
    try:
        if not can_perform_action(request.user, 'avisos', None):
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT a.*, COUNT(ae.id) as total_views
            FROM avisos a
            LEFT JOIN alert_events ae ON a.id = ae.alert_id AND ae.event_type = 'view'
            WHERE a.deleted_at IS NULL
            GROUP BY a.id
            ORDER BY a.created_at DESC
        ''')

        avisos = []
        for row in cursor.fetchall():
            avisos.append({
                'id': row['id'],
                'title': sanitizar_texto(row['title']) if row['title'] else None,
                'body_markdown': row['body_markdown'],
                'body_html': sanitizar_html(row['body_html']),
                'start_at': row['start_at'],
                'end_at': row['end_at'],
                'is_shouting': bool(row['is_shouting']),
                'has_countdown': bool(row['has_countdown']),
                'active': bool(row['active']),
                'redirect_link': row['redirect_link'],
                'image_path': row['image_path'],
                'image_size': row['image_size'],
                'created_at': row['created_at'],
                'stats': {'view': row['total_views'] or 0}
            })

        conn.close()
        return jsonify(avisos)

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/alerts', methods=['POST'])
@token_required
def create_alert():
    try:
        if not can_perform_action(request.user, 'avisos', None):
            return jsonify({'error': 'Acesso negado'}), 403

        data = request.get_json()

        if not data.get('body_markdown'):
            return jsonify({'error': 'Conteúdo é obrigatório'}), 400

        agora = datetime.now()
        start_at = data.get('start_at')
        end_at = data.get('end_at')
        active = data.get('active', False)

        if not start_at or active:
            start_at = agora.isoformat()
            if not end_at:
                end_at = (agora + timedelta(days=7)).isoformat()
            active = True

        if not end_at:
            end_at = (agora + timedelta(days=7)).isoformat()

        body_markdown = data['body_markdown']
        body_html = body_markdown.replace('**', '<strong>').replace('*', '<em>').replace('\n', '<br>').replace('\\n', '<br>').replace('\\', '')
        body_html = sanitizar_html(body_html)

        if data.get('is_shouting'):
            body_html = body_html.upper()
            body_markdown = body_markdown.upper()

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO avisos (
                title, body_markdown, body_html, start_at, end_at,
                is_shouting, has_countdown, active, redirect_link,
                image_path, image_size
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            sanitizar_texto(data.get('title')) if data.get('title') else None,
            body_markdown,
            body_html,
            start_at,
            end_at,
            1 if data.get('is_shouting') else 0,
            1 if data.get('has_countdown') else 0,
            1 if active else 0,
            data.get('redirect_link'),
            data.get('image_path'),
            data.get('image_size', 'medium')
        ))

        conn.commit()
        alert_id = cursor.lastrowid
        conn.close()

        log_admin_action(request.user, 'Criar Aviso',
                         f'Aviso "{data.get("title", "Sem título")}" criado{" (ativo imediatamente)" if active else ""}')

        return jsonify({
            'success': True,
            'id': alert_id,
            'active': active,
            'message': 'Aviso criado com sucesso' + (' e ativado imediatamente' if active else '')
        })

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/alerts/<int:alert_id>', methods=['PUT'])
@token_required
def update_alert(alert_id):
    try:
        if not can_perform_action(request.user, 'avisos', None):
            return jsonify({'error': 'Acesso negado'}), 403

        data = request.get_json()

        conn = get_db()
        cursor = conn.cursor()

        updates = []
        values = []

        if 'title' in data:
            updates.append('title = ?')
            values.append(sanitizar_texto(data['title']))

        if 'body_markdown' in data:
            body_markdown = data['body_markdown']
            body_html = body_markdown.replace('**', '<strong>').replace('*', '<em>').replace('\n', '<br>').replace('\\n', '<br>').replace('\\', '')
            body_html = sanitizar_html(body_html)

            if data.get('is_shouting'):
                body_html = body_html.upper()
                body_markdown = body_markdown.upper()

            updates.append('body_markdown = ?')
            values.append(body_markdown)
            updates.append('body_html = ?')
            values.append(body_html)

        if 'start_at' in data:
            updates.append('start_at = ?')
            values.append(data['start_at'])
        if 'end_at' in data:
            updates.append('end_at = ?')
            values.append(data['end_at'])
        if 'is_shouting' in data:
            updates.append('is_shouting = ?')
            values.append(1 if data['is_shouting'] else 0)
        if 'has_countdown' in data:
            updates.append('has_countdown = ?')
            values.append(1 if data['has_countdown'] else 0)
        if 'active' in data:
            updates.append('active = ?')
            values.append(1 if data['active'] else 0)
        if 'redirect_link' in data:
            updates.append('redirect_link = ?')
            values.append(data['redirect_link'])
        if 'image_path' in data:
            updates.append('image_path = ?')
            values.append(data['image_path'])
        if 'image_size' in data:
            updates.append('image_size = ?')
            values.append(data['image_size'])

        updates.append('updated_at = CURRENT_TIMESTAMP')
        values.append(alert_id)

        if updates:
            cursor.execute(f'''
                UPDATE avisos SET {', '.join(updates)} WHERE id = ?
            ''', values)

        conn.commit()
        conn.close()
        log_admin_action(request.user, 'Atualizar Aviso', f'Aviso ID {alert_id} atualizado')
        return jsonify({'success': True, 'message': 'Aviso atualizado com sucesso'})

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/alerts/<int:alert_id>', methods=['DELETE'])
@token_required
def delete_alert(alert_id):
    try:
        if not can_perform_action(request.user, 'avisos', None):
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('UPDATE avisos SET deleted_at = CURRENT_TIMESTAMP WHERE id = ?', (alert_id,))
        conn.commit()
        conn.close()

        log_admin_action(request.user, 'Excluir Aviso', f'Aviso ID {alert_id} excluído')

        return jsonify({'success': True, 'message': 'Aviso excluído com sucesso'})

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/alerts/upload-image', methods=['POST'])
@token_required
def upload_alert_image():
    try:
        if not can_perform_action(request.user, 'upload', None):
            return jsonify({'error': 'Acesso negado'}), 403

        if 'image' not in request.files:
            return jsonify({'error': 'Nenhuma imagem enviada'}), 400

        file = request.files['image']

        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo vazio'}), 400

        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''

        if ext not in allowed_extensions:
            return jsonify({'error': 'Tipo de arquivo não permitido. Use: png, jpg, jpeg, gif, webp'}), 400

        file_content = file.read(1024)
        file.seek(0)

        magic_types = {
            b'\xff\xd8\xff': 'jpg',
            b'\x89PNG\r\n\x1a\n': 'png',
            b'GIF87a': 'gif',
            b'GIF89a': 'gif',
            b'RIFF': 'webp'
        }

        detected_type = None
        for magic, file_type in magic_types.items():
            if file_content.startswith(magic):
                detected_type = file_type
                break

        if not detected_type or detected_type != ext:
            return jsonify({'error': 'Arquivo corrompido ou tipo inválido'}), 400

        if not os.path.exists(UPLOAD_DIR):
            os.makedirs(UPLOAD_DIR)
        if not os.path.exists(THUMBNAIL_DIR):
            os.makedirs(THUMBNAIL_DIR)

        filename = f"alert_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(8)}.{ext}"
        filepath = os.path.join('uploads', filename)
        full_path = os.path.join(BASE_DIR, filepath)

        file.save(full_path)

        if os.path.getsize(full_path) > 5 * 1024 * 1024:
            os.remove(full_path)
            return jsonify({'error': 'Imagem muito grande. Máximo 5MB'}), 400

        thumbnail_url = generate_thumbnail(full_path, filename, f'image/{ext}')

        return jsonify({
            'success': True,
            'filepath': f'/{filepath}',
            'thumbnail_url': thumbnail_url,
            'message': 'Imagem enviada com sucesso'
        })

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ROTAS DE BACKUP
# ============================================

@app.route('/api/admin/backup', methods=['GET'])
@token_required
def get_backup():
    try:
        if not can_perform_action(request.user, 'backup', None):
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT t.nome as turma, a.edv, a.nome
            FROM alunos a
            JOIN turmas t ON a.turma_id = t.id
            WHERE a.is_admin = 0
            ORDER BY t.nome, a.nome
        ''')

        alunos = {}
        for row in cursor.fetchall():
            if row['turma'] not in alunos:
                alunos[row['turma']] = {}
            alunos[row['turma']][row['edv']] = row['nome']

        cursor.execute('''
            SELECT t.nome as turma, e.semana_numero, e.data_inicio, e.data_fim, e.dupla
            FROM escalas e
            JOIN turmas t ON e.turma_id = t.id
            ORDER BY t.nome, e.semana_numero
        ''')

        escalas = {}
        for row in cursor.fetchall():
            if row['turma'] not in escalas:
                escalas[row['turma']] = []
            escalas[row['turma']].append({
                'semana': row['semana_numero'],
                'data_inicio': row['data_inicio'],
                'data_fim': row['data_fim'],
                'dupla': json.loads(row['dupla'])
            })

        conn.close()

        return jsonify({
            'alunos': alunos,
            'escalas': escalas,
            'exportado_em': datetime.now().isoformat()
        })

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/backup', methods=['POST'])
@token_required
def restore_backup():
    try:
        if not can_perform_action(request.user, 'backup', None):
            return jsonify({'error': 'Acesso negado'}), 403

        data = request.get_json()

        if not data or 'alunos' not in data or 'escalas' not in data:
            return jsonify({'error': 'Dados de backup inválidos'}), 400

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('DELETE FROM escalas')
        cursor.execute('DELETE FROM alunos')

        for turma_nome, alunos_dict in data['alunos'].items():
            turma_nome = sanitizar_texto(turma_nome)
            cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
            turma = cursor.fetchone()
            if turma:
                for edv, nome in alunos_dict.items():
                    edv = sanitizar_texto(edv)
                    nome = sanitizar_texto(nome)
                    if validar_edv(edv) and validar_nome(nome):
                        is_admin = 1 if '👩‍🏫' in nome or '👨‍🏫' in nome else 0
                        cursor.execute('''
                            INSERT OR IGNORE INTO alunos (edv, nome, turma_id, is_admin)
                            VALUES (?, ?, ?, ?)
                        ''', (edv, nome, turma['id'], is_admin))

        for turma_nome, escalas_list in data['escalas'].items():
            turma_nome = sanitizar_texto(turma_nome)
            cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
            turma = cursor.fetchone()
            if turma:
                for escala in escalas_list:
                    cursor.execute('''
                        INSERT OR IGNORE INTO escalas (turma_id, semana_numero, data_inicio, data_fim, dupla)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (
                        turma['id'],
                        escala['semana'],
                        escala['data_inicio'],
                        escala['data_fim'],
                        json.dumps(escala['dupla'])
                    ))

        conn.commit()
        conn.close()

        log_admin_action(request.user, 'Restaurar Backup', 'Backup restaurado com sucesso')

        return jsonify({'success': True, 'message': 'Backup restaurado com sucesso'})

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ROTAS DE UPLOADS E MINIATURAS
# ============================================

@app.route('/api/admin/uploads/list', methods=['GET'])
@token_required
def list_uploads():
    try:
        if not can_perform_action(request.user, 'upload', None):
            return jsonify({'error': 'Acesso negado'}), 403

        if not os.path.exists(UPLOAD_DIR):
            return jsonify({'files': []})

        files = []
        for filename in os.listdir(UPLOAD_DIR):
            if not validate_file_path(UPLOAD_DIR, filename):
                continue

            filepath = os.path.join(UPLOAD_DIR, filename)
            if not os.path.isfile(filepath):
                continue

            try:
                mime_type = magic.from_file(filepath, mime=True)
            except:
                mime_type = mimetypes.guess_type(filename)[0] or 'application/octet-stream'

            ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
            allowed_extensions = ['png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf', 'txt', 'csv', 'xlsx', 'xls', 'doc', 'docx']

            if ext not in allowed_extensions:
                continue

            thumbnail_url = None
            if mime_type.startswith('image/'):
                thumbnail_url = generate_thumbnail(filepath, filename, mime_type)

            icon = None
            if not thumbnail_url:
                icon = get_file_icon(mime_type, filename)

            file_stat = os.stat(filepath)

            files.append({
                'name': filename,
                'url': f'/uploads/{filename}',
                'thumbnail_url': thumbnail_url,
                'icon': icon,
                'type': mime_type,
                'size': file_stat.st_size,
                'size_formatted': f"{file_stat.st_size / 1024:.1f} KB" if file_stat.st_size < 1024 * 1024 else f"{file_stat.st_size / (1024 * 1024):.1f} MB",
                'modified': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
                'modified_formatted': datetime.fromtimestamp(file_stat.st_mtime).strftime('%d/%m/%Y %H:%M')
            })

        files.sort(key=lambda x: x['modified'], reverse=True)

        return jsonify({
            'files': files,
            'total': len(files),
            'thumbnails_enabled': True
        })

    except Exception as e:
        print(f"Erro ao listar uploads: {e}")
        return jsonify({'error': 'Erro interno ao listar arquivos'}), 500

@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    filename = sanitizar_texto(filename)

    if not validate_file_path(UPLOAD_DIR, filename):
        return jsonify({'error': 'Acesso negado'}), 403

    filepath = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'Arquivo não encontrado'}), 404

    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext not in ['png', 'jpg', 'jpeg', 'gif', 'webp']:
        return jsonify({'error': 'Tipo de arquivo não permitido'}), 403

    return send_from_directory(UPLOAD_DIR, filename)

@app.route('/thumbnails/<path:filename>')
def serve_thumbnail(filename):
    try:
        filename = sanitizar_texto(filename)

        if not validate_file_path(THUMBNAIL_DIR, filename):
            return jsonify({'error': 'Acesso negado'}), 403

        filepath = os.path.join(THUMBNAIL_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({'error': 'Miniatura não encontrada'}), 404

        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        if ext not in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
            return jsonify({'error': 'Tipo de arquivo não permitido'}), 403

        return send_from_directory(THUMBNAIL_DIR, filename)

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ROTAS DE HISTÓRICO
# ============================================

@app.route('/api/admin/historico/recente', methods=['GET'])
@token_required
def get_recent_history():
    try:
        if not can_perform_action(request.user, 'historico', None):
            return jsonify({'error': 'Acesso negado'}), 403

        dias = request.args.get('dias', 5, type=int)

        if dias < 1 or dias > 30:
            dias = 5

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT id, usuario, acao, detalhes, turma, item_afetado,
                   datetime(data_hora, '-3 hours') as data_hora_brasil
            FROM historico_acoes
            WHERE data_hora >= datetime('now', ?)
            ORDER BY data_hora DESC
            LIMIT 100
        ''', (f'-{dias} days',))

        historico = []
        for row in cursor.fetchall():
            historico.append({
                'id': row['id'],
                'usuario': row['usuario'],
                'acao': row['acao'],
                'detalhes': row['detalhes'],
                'turma': row['turma'],
                'item_afetado': row['item_afetado'],
                'data_hora': row['data_hora_brasil']
            })

        conn.close()
        return jsonify(historico)

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

@app.route('/api/admin/historico', methods=['DELETE'])
@token_required
def clear_history():
    try:
        if get_admin_type(request.user) != 'global':
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT COUNT(*) as total FROM historico_acoes')
        total = cursor.fetchone()['total']

        cursor.execute('DELETE FROM historico_acoes')
        conn.commit()
        conn.close()

        log_admin_action(request.user, 'Limpar Histórico', f'{total} registros removidos do histórico')

        return jsonify({
            'success': True,
            'removidos': total,
            'message': f'{total} registros removidos do histórico'
        })

    except Exception as e:
        return jsonify({'error': 'Erro interno'}), 500

# ============================================
# ROTAS DE TESTE
# ============================================

@app.route('/api/historico/teste', methods=['GET'])
def test_history():
    return jsonify([
        {
            'id': 1,
            'data_hora': datetime.now().isoformat(),
            'usuario': 'Admin Teste',
            'acao': 'Login',
            'detalhes': 'Usuário fez login no sistema',
            'turma': TURMAS['formare'],
            'item_afetado': None
        }
    ])

# ============================================
# ROTAS DE RELATÓRIO EXCEL
# ============================================

@app.route('/api/presenca/relatorio_excel/<mes>', methods=['GET'])
@token_required
def generate_excel_report(mes):
    """Gera relatório Excel com cabeçalho da turma, contadores por aluno e filtros automáticos"""
    try:
        turma_nome = request.args.get('turma', TURMAS['formare'])
        turma_nome = sanitizar_texto(turma_nome)

        if not can_perform_action(request.user, 'relatorios', turma_nome):
            return jsonify({'error': 'Acesso negado'}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM turmas WHERE nome = ?', (turma_nome,))
        turma = cursor.fetchone()

        if not turma:
            conn.close()
            return jsonify({'error': 'Turma não encontrada'}), 404

        cursor.execute('''
            SELECT data, registros FROM presenca
            WHERE turma_id = ? AND strftime('%Y-%m', data) = ?
            ORDER BY data
        ''', (turma['id'], mes))

        registros = cursor.fetchall()

        cursor.execute('''
            SELECT nome FROM alunos
            WHERE turma_id = ? AND is_admin = 0
            ORDER BY nome
        ''', (turma['id'],))
        alunos = [row['nome'] for row in cursor.fetchall()]

        conn.close()

        if not registros:
            return jsonify({'error': 'Nenhum registro encontrado para este mês'}), 404

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Chamadas {mes}'

        # ============================================
        # CÁLCULO DAS COLUNAS
        # Coluna A = Nome
        # Colunas B até N = Datas (onde N = 1 + len(datas))
        # Coluna N+1 = Presenças
        # Coluna N+2 = Ausências
        # Coluna N+3 = Atestados
        # ============================================

        num_datas = len(registros)
        col_presencas = num_datas + 2  # Após as datas
        col_ausencias = num_datas + 3
        col_atestados = num_datas + 4

        # Última coluna para o título
        ultima_coluna_letra = openpyxl.utils.get_column_letter(col_atestados)

        # ============================================
        # TÍTULO (Linha 1)
        # ============================================
        ws.merge_cells(f'A1:{ultima_coluna_letra}1')

        titulo_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        titulo_font = Font(color="FFFFFF", bold=True, size=14)

        titulo_cell = ws['A1']
        titulo_cell.value = f"📊 RELATÓRIO DE FREQUÊNCIA - {turma_nome.upper()}"
        titulo_cell.font = titulo_font
        titulo_cell.fill = titulo_fill
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[1].height = 30

        # ============================================
        # CABEÇALHO (Linha 3) - FILTROS FICARÃO AQUI!
        # ============================================
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Coluna Nome
        ws['A3'] = 'Nome do Aluno'
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

        # Colunas de Datas
        datas = []
        for row in registros:
            data_obj = datetime.strptime(row['data'], '%Y-%m-%d')
            data_str_formatada = data_obj.strftime('%d/%m')
            datas.append({
                'obj': data_obj,
                'str': data_str_formatada,
                'full': row['data']
            })

        datas.sort(key=lambda x: x['obj'])

        for col_idx, data in enumerate(datas, start=2):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = data['str']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Colunas de Contadores
        ws.cell(row=3, column=col_presencas, value='PRESENÇAS')
        ws.cell(row=3, column=col_presencas).font = header_font
        ws.cell(row=3, column=col_presencas).fill = header_fill
        ws.cell(row=3, column=col_presencas).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=3, column=col_ausencias, value='AUSÊNCIAS')
        ws.cell(row=3, column=col_ausencias).font = header_font
        ws.cell(row=3, column=col_ausencias).fill = header_fill
        ws.cell(row=3, column=col_ausencias).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=3, column=col_atestados, value='ATESTADOS')
        ws.cell(row=3, column=col_atestados).font = header_font
        ws.cell(row=3, column=col_atestados).fill = header_fill
        ws.cell(row=3, column=col_atestados).alignment = Alignment(horizontal='center', vertical='center')

        # ============================================
        # CORES PARA STATUS DAS CÉLULAS
        # ============================================
        present_fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")
        absent_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
        atestado_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        empty_fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")

        # ============================================
        # PREENCHER DADOS DOS ALUNOS (a partir da Linha 4)
        # ============================================
        for row_idx, aluno in enumerate(alunos, start=4):
            # Nome do aluno
            nome_cell = ws.cell(row=row_idx, column=1, value=aluno)
            nome_cell.font = Font(bold=True, size=11)
            nome_cell.alignment = Alignment(horizontal='left', vertical='center')

            # Mapear status por data
            registros_por_aluno = {}
            for registro in registros:
                for r in json.loads(registro['registros']):
                    if r['name'] == aluno:
                        registros_por_aluno[registro['data']] = r['status']

            # Contadores
            total_presencas = 0
            total_ausencias = 0
            total_atestados = 0

            # Preencher cada data
            for col_idx, data in enumerate(datas, start=2):
                status = registros_por_aluno.get(data['full'], '')
                cell = ws.cell(row=row_idx, column=col_idx, value=status)
                cell.alignment = Alignment(horizontal='center', vertical='center')

                if status == 'PRESENTE':
                    cell.fill = present_fill
                    cell.font = white_font
                    total_presencas += 1
                elif status == 'AUSENTE':
                    cell.fill = absent_fill
                    cell.font = white_font
                    total_ausencias += 1
                elif status == 'ATESTADO':
                    cell.fill = atestado_fill
                    cell.font = white_font
                    total_atestados += 1
                elif status == '':
                    cell.fill = empty_fill
                    cell.value = '-'

            # Preencher contadores no final da linha
            presencas_cell = ws.cell(row=row_idx, column=col_presencas, value=total_presencas)
            presencas_cell.alignment = Alignment(horizontal='center', vertical='center')
            presencas_cell.font = Font(bold=True, size=11)

            ausencias_cell = ws.cell(row=row_idx, column=col_ausencias, value=total_ausencias)
            ausencias_cell.alignment = Alignment(horizontal='center', vertical='center')
            ausencias_cell.font = Font(bold=True, size=11)

            atestados_cell = ws.cell(row=row_idx, column=col_atestados, value=total_atestados)
            atestados_cell.alignment = Alignment(horizontal='center', vertical='center')
            atestados_cell.font = Font(bold=True, size=11)

        # ============================================
        # AJUSTAR LARGURA DAS COLUNAS
        # ============================================
        ws.column_dimensions['A'].width = 35

        for col_idx in range(2, len(datas) + 2):
            col_letter = ws.cell(row=3, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 12

        ws.column_dimensions[openpyxl.utils.get_column_letter(col_presencas)].width = 12
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_ausencias)].width = 12
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_atestados)].width = 12

        # ============================================
        # BORDAS
        # ============================================
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=3, max_row=len(alunos) + 3, min_col=1, max_col=col_atestados):
            for cell in row:
                cell.border = thin_border

        # ============================================
        # FILTROS AUTOMÁTICOS (NA LINHA 3, ONDE ESTÃO OS CABEÇALHOS!)
        # ============================================
        # Define o intervalo exato onde os filtros serão aplicados
        # Da coluna A (1) até a última coluna (col_atestados)
        # Da linha 3 (cabeçalho) até a última linha com dados
        ultima_linha = len(alunos) + 3
        ultima_coluna_letra_filter = openpyxl.utils.get_column_letter(col_atestados)

        # Aplica filtro no intervalo correto
        ws.auto_filter.ref = f"A3:{ultima_coluna_letra_filter}{ultima_linha}"

        # ============================================
        # NOME DO ARQUIVO
        # ============================================
        def sanitizar_nome_arquivo(nome):
            nome = nome.replace(' ', '_')
            nome = nome.replace('(', '')
            nome = nome.replace(')', '')
            nome = nome.replace('+', 'mais')
            nome = nome.replace('-', '_')
            nome = nome.replace('á', 'a').replace('ã', 'a').replace('â', 'a')
            nome = nome.replace('é', 'e').replace('ê', 'e')
            nome = nome.replace('í', 'i').replace('î', 'i')
            nome = nome.replace('ó', 'o').replace('ô', 'o').replace('õ', 'o')
            nome = nome.replace('ú', 'u').replace('û', 'u')
            nome = nome.replace('ç', 'c')
            nome = nome.replace('__', '_').replace('___', '_')
            return nome

        nome_arquivo_turma = sanitizar_nome_arquivo(turma_nome)

        meses_nomes = {
            '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
            '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
            '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
        }

        ano_mes = mes.split('-')
        ano = ano_mes[0] if len(ano_mes) > 0 else '2026'
        mes_num = ano_mes[1] if len(ano_mes) > 1 else '04'

        nome_mes = meses_nomes.get(mes_num, mes_num)

        nome_arquivo = f'relatorio_chamadas_{nome_arquivo_turma}_{nome_mes}_{ano}.xlsx'

        print(f"📊 Gerando relatório para turma: {turma_nome}")
        print(f"📁 Nome do arquivo: {nome_arquivo}")
        print(f"📌 Filtros aplicados no intervalo: A3:{ultima_coluna_letra_filter}{ultima_linha}")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        print(f"Erro ao gerar relatório Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Erro interno ao gerar relatório'}), 500
# ============================================
# ROTAS SUPER ADMIN (GERENCIAMENTO DE ADMINS)
# ============================================

@app.route('/api/super/admins', methods=['GET'])
@token_required
def super_list_admins():
    """Lista todos os administradores (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    admins = get_all_admins()
    return jsonify({
        'success': True,
        'admins': [dict(admin) for admin in admins]
    })

@app.route('/api/super/admins', methods=['POST'])
@token_required
def super_create_admin():
    """Cria um novo administrador (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    admin_type = data.get('admin_type')
    turma_permitida = data.get('turma_permitida')

    if not username or not password or not admin_type:
        return jsonify({'error': 'Usuário, senha e tipo são obrigatórios'}), 400

    if admin_type not in ['global', 'formare', 'turma']:
        return jsonify({'error': 'Tipo de admin inválido'}), 400

    if admin_type == 'turma' and not turma_permitida:
        return jsonify({'error': 'Admin de turma precisa de uma turma permitida'}), 400

    existing = get_admin_by_username(username)
    if existing:
        return jsonify({'error': 'Usuário já existe'}), 409

    create_admin(username, password, admin_type, turma_permitida, request.user)
    log_admin_action_db(request.user, 'create_admin', f'Criou admin: {username} (tipo: {admin_type})')

    return jsonify({'success': True, 'message': f'Admin {username} criado com sucesso'})

@app.route('/api/super/admins/<username>', methods=['PUT'])
@token_required
def super_update_admin(username):
    """Atualiza senha de um administrador (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    data = request.get_json()
    new_password = data.get('password')

    if not new_password:
        return jsonify({'error': 'Nova senha é obrigatória'}), 400

    if update_admin_password(username, new_password):
        log_admin_action_db(request.user, 'update_admin_password', f'Alterou senha do admin: {username}')
        return jsonify({'success': True, 'message': f'Senha de {username} atualizada'})
    else:
        return jsonify({'error': 'Admin não encontrado'}), 404

@app.route('/api/super/admins/<username>/rename', methods=['PUT'])
@token_required
def super_rename_admin(username):
    """Renomeia um administrador (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    if username == request.user:
        return jsonify({'error': 'Não é possível renomear seu próprio usuário'}), 400

    data = request.get_json()
    new_username = data.get('new_username')

    if not new_username:
        return jsonify({'error': 'Novo nome de usuário é obrigatório'}), 400

    conn = get_admins_db()
    cursor = conn.cursor()

    cursor.execute('SELECT id FROM admins WHERE username = ?', (new_username,))
    if cursor.fetchone():
        conn.close()
        return jsonify({'error': f'Usuário "{new_username}" já existe'}), 409

    cursor.execute('UPDATE admins SET username = ?, updated_at = CURRENT_TIMESTAMP WHERE username = ?', (new_username, username))

    if cursor.rowcount == 0:
        conn.close()
        return jsonify({'error': 'Admin não encontrado'}), 404

    conn.commit()
    conn.close()

    log_admin_action_db(request.user, 'rename_admin', f'Renomeou admin: {username} -> {new_username}')

    return jsonify({'success': True, 'message': f'Admin renomeado de "{username}" para "{new_username}"'})

@app.route('/api/super/admins/<username>', methods=['DELETE'])
@token_required
def super_delete_admin(username):
    """Remove um administrador (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    if username == request.user:
        return jsonify({'error': 'Não é possível deletar seu próprio usuário'}), 400

    if delete_admin(username):
        log_admin_action_db(request.user, 'delete_admin', f'Removeu admin: {username}')
        return jsonify({'success': True, 'message': f'Admin {username} removido'})
    else:
        return jsonify({'error': 'Admin não encontrado'}), 404

@app.route('/api/super/blocked-ips', methods=['GET'])
@token_required
def super_list_blocked_ips():
    """Lista IPs bloqueados (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    ips = get_blocked_ips()
    return jsonify({
        'success': True,
        'blocked_ips': [dict(ip) for ip in ips]
    })

@app.route('/api/super/blocked-ips', methods=['POST'])
@token_required
def super_block_ip():
    """Bloqueia um IP (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    data = request.get_json()
    ip = data.get('ip')
    reason = data.get('reason', 'Bloqueado manualmente')
    expires_minutes = data.get('expires_minutes', 60)

    if not ip:
        return jsonify({'error': 'IP é obrigatório'}), 400

    add_blocked_ip(ip, reason, request.user, expires_minutes)
    log_admin_action_db(request.user, 'block_ip', f'Bloqueou IP: {ip} - Motivo: {reason}')

    return jsonify({'success': True, 'message': f'IP {ip} bloqueado por {expires_minutes} minutos'})

@app.route('/api/super/blocked-ips/<ip>', methods=['DELETE'])
@token_required
def super_unblock_ip(ip):
    """Desbloqueia um IP (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    if remove_blocked_ip(ip):
        log_admin_action_db(request.user, 'unblock_ip', f'Desbloqueou IP: {ip}')
        return jsonify({'success': True, 'message': f'IP {ip} desbloqueado'})
    else:
        return jsonify({'error': 'IP não encontrado'}), 404

@app.route('/api/super/admin-logs', methods=['GET'])
@token_required
def super_get_admin_logs():
    """Retorna logs de ações dos admins (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    limit = request.args.get('limit', 100, type=int)

    conn = get_admins_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM admin_logs ORDER BY created_at DESC LIMIT ?', (limit,))
    logs = cursor.fetchall()
    conn.close()

    return jsonify({'success': True, 'logs': [dict(log) for log in logs]})

@app.route('/api/super/db-stats', methods=['GET'])
@token_required
def super_db_stats():
    """Retorna estatísticas do banco de dados (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    db_size = os.path.getsize(DATABASE) if os.path.exists(DATABASE) else 0
    admins_db_size = os.path.getsize(ADMINS_DB) if os.path.exists(ADMINS_DB) else 0

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM alunos WHERE is_admin = 0")
    total_alunos = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM turmas")
    total_turmas = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM presenca")
    total_presencas = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM avisos WHERE deleted_at IS NULL")
    total_avisos = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM historico_acoes")
    total_logs = cursor.fetchone()[0]

    conn.close()

    conn_admins = get_admins_db()
    cursor_admins = conn_admins.cursor()

    cursor_admins.execute("SELECT COUNT(*) FROM admins")
    total_admins = cursor_admins.fetchone()[0]

    cursor_admins.execute("SELECT COUNT(*) FROM admins WHERE admin_type = 'global'")
    total_globals = cursor_admins.fetchone()[0]

    cursor_admins.execute("SELECT COUNT(*) FROM admins WHERE admin_type = 'formare'")
    total_formare = cursor_admins.fetchone()[0]

    cursor_admins.execute("SELECT COUNT(*) FROM admins WHERE admin_type = 'turma'")
    total_turma = cursor_admins.fetchone()[0]

    cursor_admins.execute("SELECT COUNT(*) FROM blocked_ips WHERE expires_at IS NULL OR expires_at > datetime('now')")
    total_blocked_ips = cursor_admins.fetchone()[0]

    conn_admins.close()

    return jsonify({
        'success': True,
        'stats': {
            'database_size_mb': round(db_size / (1024 * 1024), 2),
            'admins_db_size_mb': round(admins_db_size / (1024 * 1024), 2),
            'total_alunos': total_alunos,
            'total_turmas': total_turmas,
            'total_presencas': total_presencas,
            'total_avisos': total_avisos,
            'total_logs': total_logs,
            'total_admins': total_admins,
            'total_globals': total_globals,
            'total_formare': total_formare,
            'total_turma': total_turma,
            'total_blocked_ips': total_blocked_ips
        }
    })

@app.route('/api/super/clear-all-data', methods=['POST'])
@token_required
def super_clear_all_data():
    """Limpa todos os dados do sistema (apenas super admin) - requer confirmação"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    data = request.get_json()
    confirmation = data.get('confirmation')

    if confirmation != 'CONFIRMAR_LIMPEZA_TOTAL':
        return jsonify({'error': 'Confirmação inválida. Use "CONFIRMAR_LIMPEZA_TOTAL"'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM alunos")
    cursor.execute("DELETE FROM escalas")
    cursor.execute("DELETE FROM escalas_historico")
    cursor.execute("DELETE FROM presenca")
    cursor.execute("DELETE FROM avisos")
    cursor.execute("DELETE FROM alert_events")
    cursor.execute("DELETE FROM historico_acoes")
    cursor.execute("DELETE FROM auto_generation_log")

    for turma_nome in TURMAS.values():
        cursor.execute('INSERT OR IGNORE INTO turmas (nome) VALUES (?)', (turma_nome,))

    conn.commit()
    conn.close()

    log_admin_action_db(request.user, 'clear_all_data', 'Limpeza total de dados realizada')

    return jsonify({'success': True, 'message': 'Todos os dados foram limpos com sucesso'})

@app.route('/api/super/export-logs-excel', methods=['GET'])
@token_required
def super_export_logs_excel():
    """Exporta todos os logs em Excel (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, usuario, acao, detalhes, turma, item_afetado,
               datetime(data_hora, '-3 hours') as data_hora_brasil
        FROM historico_acoes ORDER BY data_hora DESC
    ''')
    logs = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Logs do Sistema'

    headers = ['ID', 'Data/Hora', 'Usuário', 'Ação', 'Detalhes', 'Turma', 'Item Afetado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=log['id'])
        ws.cell(row=row_idx, column=2, value=log['data_hora_brasil'])
        ws.cell(row=row_idx, column=3, value=log['usuario'])
        ws.cell(row=row_idx, column=4, value=log['acao'])
        ws.cell(row=row_idx, column=5, value=log['detalhes'] or '')
        ws.cell(row=row_idx, column=6, value=log['turma'] or '')
        ws.cell(row=row_idx, column=7, value=log['item_afetado'] or '')

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'logs_completos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

# ============================================
# BACKUP COMPLETO
# ============================================

@app.route('/api/super/backup-full', methods=['GET'])
@token_required
def super_backup_full():
    """Exporta backup completo do sistema (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    backup_data = {'exportado_em': datetime.now().isoformat(), 'exportado_por': request.user, 'versao': '1.0', 'dados': {}}

    conn = get_db()
    cursor = conn.cursor()

    tabelas = ['turmas', 'alunos', 'escalas', 'escalas_historico', 'avisos', 'alert_events', 'presenca', 'historico_acoes', 'auto_generation_log']
    for tabela in tabelas:
        try:
            cursor.execute(f"SELECT * FROM {tabela}")
            backup_data['dados'][tabela] = [dict(row) for row in cursor.fetchall()]
        except:
            backup_data['dados'][tabela] = []
    conn.close()

    conn_admins = get_admins_db()
    cursor_admins = conn_admins.cursor()

    try:
        cursor_admins.execute("SELECT id, username, password_hash, admin_type, turma_permitida, created_by, created_at, updated_at FROM admins")
        backup_data['dados']['admins'] = [dict(row) for row in cursor_admins.fetchall()]
    except:
        backup_data['dados']['admins'] = []

    try:
        cursor_admins.execute("SELECT * FROM blocked_ips")
        backup_data['dados']['blocked_ips'] = [dict(row) for row in cursor_admins.fetchall()]
    except:
        backup_data['dados']['blocked_ips'] = []

    conn_admins.close()

    return jsonify(backup_data)

@app.route('/api/super/backup-restore', methods=['POST'])
@token_required
def super_backup_restore():
    """Restaura backup completo do sistema (apenas super admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    data = request.get_json()
    if not data or 'dados' not in data:
        return jsonify({'error': 'Dados de backup inválidos'}), 400

    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM auto_generation_log")
        cursor.execute("DELETE FROM historico_acoes")
        cursor.execute("DELETE FROM presenca")
        cursor.execute("DELETE FROM alert_events")
        cursor.execute("DELETE FROM avisos")
        cursor.execute("DELETE FROM escalas_historico")
        cursor.execute("DELETE FROM escalas")
        cursor.execute("DELETE FROM alunos")

        for turma in data['dados'].get('turmas', []):
            cursor.execute('INSERT OR IGNORE INTO turmas (id, nome, created_at) VALUES (?, ?, ?)', (turma.get('id'), turma.get('nome'), turma.get('created_at')))

        for aluno in data['dados'].get('alunos', []):
            cursor.execute('INSERT OR IGNORE INTO alunos (id, edv, nome, turma_id, is_admin, created_at) VALUES (?, ?, ?, ?, ?, ?)', (aluno.get('id'), aluno.get('edv'), aluno.get('nome'), aluno.get('turma_id'), aluno.get('is_admin', 0), aluno.get('created_at')))

        for escala in data['dados'].get('escalas', []):
            cursor.execute('INSERT OR IGNORE INTO escalas (id, turma_id, semana_numero, data_inicio, data_fim, dupla, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)', (escala.get('id'), escala.get('turma_id'), escala.get('semana_numero'), escala.get('data_inicio'), escala.get('data_fim'), escala.get('dupla'), escala.get('created_at')))

        for aviso in data['dados'].get('avisos', []):
            cursor.execute('INSERT OR IGNORE INTO avisos (id, title, body_markdown, body_html, start_at, end_at, is_shouting, has_countdown, active, redirect_link, image_path, image_size, deleted_at, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', (aviso.get('id'), aviso.get('title'), aviso.get('body_markdown'), aviso.get('body_html'), aviso.get('start_at'), aviso.get('end_at'), aviso.get('is_shouting', 0), aviso.get('has_countdown', 0), aviso.get('active', 0), aviso.get('redirect_link'), aviso.get('image_path'), aviso.get('image_size', 'medium'), aviso.get('deleted_at'), aviso.get('created_at'), aviso.get('updated_at')))

        for presenca in data['dados'].get('presenca', []):
            cursor.execute('INSERT OR IGNORE INTO presenca (id, data, turma_id, registros, responsavel, created_by, updated_by, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)', (presenca.get('id'), presenca.get('data'), presenca.get('turma_id'), presenca.get('registros'), presenca.get('responsavel'), presenca.get('created_by'), presenca.get('updated_by'), presenca.get('created_at'), presenca.get('updated_at')))

        for log in data['dados'].get('historico_acoes', []):
            cursor.execute('INSERT OR IGNORE INTO historico_acoes (id, usuario, acao, detalhes, turma, item_afetado, data_hora) VALUES (?, ?, ?, ?, ?, ?, ?)', (log.get('id'), log.get('usuario'), log.get('acao'), log.get('detalhes'), log.get('turma'), log.get('item_afetado'), log.get('data_hora')))

        conn.commit()
        conn.close()

        if 'admins' in data['dados'] and data['dados']['admins']:
            conn_admins = get_admins_db()
            cursor_admins = conn_admins.cursor()
            current_admin = request.user
            for admin in data['dados']['admins']:
                if admin.get('username') != current_admin:
                    cursor_admins.execute('INSERT OR REPLACE INTO admins (id, username, password_hash, admin_type, turma_permitida, created_by, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', (admin.get('id'), admin.get('username'), admin.get('password_hash'), admin.get('admin_type'), admin.get('turma_permitida'), admin.get('created_by'), admin.get('created_at'), admin.get('updated_at')))
            conn_admins.commit()
            conn_admins.close()

        log_admin_action_db(request.user, 'restore_backup', 'Backup completo restaurado')
        return jsonify({'success': True, 'message': 'Backup restaurado com sucesso'})

    except Exception as e:
        return jsonify({'error': f'Erro ao restaurar backup: {str(e)}'}), 500

# ============================================
# LOGIN COM CÓDIGO DE VERIFICAÇÃO (2FA)
# ============================================

verification_codes = {}

def send_verification_email(email_destino, codigo):
    """Envia email com código de verificação"""
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        email_from = "erickmth4565@gmail.com"
        senha = "rkknqmpsmhlzurrs"

        msg = MIMEMultipart("alternative")
        msg["From"] = email_from
        msg["To"] = email_destino
        msg["Subject"] = "Código de Verificação - Super Admin"

        html = f"""
        <!DOCTYPE html>
        <html>
        <head><meta charset="UTF-8"></head>
        <body style="margin:0; padding:0; background-color:#f4f6f8; font-family:Arial, sans-serif;">
          <table width="100%" cellpadding="0" cellspacing="0"><tr><td align="center">
            <table width="400" cellpadding="0" cellspacing="0" style="background:#ffffff; margin-top:40px; border-radius:12px; overflow:hidden; box-shadow:0 4px 20px rgba(0,0,0,0.08);">
              <tr><td style="background:linear-gradient(135deg,#4f46e5,#7c3aed); padding:20px; text-align:center;"><h1 style="color:#ffffff; margin:0;">🔐 Verificação</h1></td></tr>
              <tr><td style="padding:30px; text-align:center;">
                <p>Use o código abaixo para continuar:</p>
                <div style="margin:25px 0;"><span style="display:inline-block; padding:15px 25px; font-size:28px; letter-spacing:5px; font-weight:bold; color:#4f46e5; background:#f3f4f6; border-radius:8px;">{codigo}</span></div>
                <p><small>Este código expira em 5 minutos.</small></p>
              </td></tr>
              <tr><td style="padding:20px; text-align:center; font-size:12px; color:#999;">Se você não solicitou este código, ignore este email.</td></tr>
            </table>
          </td></tr></table>
        </body>
        </html>
        """

        msg.attach(MIMEText(html, "html"))

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(email_from, senha)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Erro ao enviar email: {e}")
        return False
@app.route('/api/super/send-code', methods=['POST'])
def send_verification_code():
    """Envia código de verificação para o email do super admin"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({'success': False, 'error': 'Usuário e senha são obrigatórios'}), 400

    admin = get_admin_by_username(username)
    if not admin or not check_password_hash(admin['password_hash'], password):
        return jsonify({'success': False, 'error': 'Credenciais inválidas'}), 401

    if not is_super_admin(username):
        return jsonify({'success': False, 'error': 'Acesso negado. Apenas Super Admin.'}), 403

    # Gerar código de 6 dígitos
    codigo = str(random.randint(100000, 999999))
    verification_codes[username] = {'code': codigo, 'expires': datetime.now() + timedelta(minutes=5)}

    # Você precisa definir o email do super admin
    # Pode buscar do banco ou usar um email fixo
    email_destino = "erickleguisamon@gmail.com"  # ← ALTERE para o email do super admin

    # Enviar email
    if send_verification_email(email_destino, codigo):
        log_admin_action_db(username, 'send_2fa_code', f'Código enviado para {email_destino}')
        return jsonify({'success': True, 'message': 'Código enviado para seu email!'})
    else:
        return jsonify({'success': False, 'error': 'Erro ao enviar email. Tente novamente.'}), 500

@app.route('/api/super/verify-code', methods=['POST'])
def verify_code():
    """Verifica o código de autenticação - MODO TESTE"""
    data = request.get_json()
    username = data.get('username')
    code = data.get('code')

    if not username or not code:
        return jsonify({'success': False, 'error': 'Usuário e código são obrigatórios'}), 400

    # MODO TESTE: Aceita o código 123456 OU qualquer código armazenado
    admin = get_admin_by_username(username)
    if not admin:
        return jsonify({'success': False, 'error': 'Usuário não encontrado'}), 404

    # Verifica se o código é 123456 (modo teste) OU o código armazenado
    stored = verification_codes.get(username)
    is_valid = (code == "123456") or (stored and datetime.now() <= stored['expires'] and stored['code'] == code)

    if not is_valid:
        return jsonify({'success': False, 'error': 'Código inválido ou expirado'}), 400

    # Limpa o código usado
    if username in verification_codes:
        del verification_codes[username]

    admin_type = admin['admin_type']
    turmas_permitidas = get_admin_turmas_permitidas(username)
    permissoes = get_admin_permissoes(username)

    payload = {
        'username': username,
        'exp': datetime.utcnow() + timedelta(hours=JWT_EXPIRATION),
        'iat': datetime.utcnow()
    }
    token = jwt.encode(payload, JWT_SECRET, algorithm='HS256')

    log_admin_action_db(username, 'login_2fa', 'Login com verificação em duas etapas (modo teste)')

    print(f"✅ Login bem-sucedido para {username} (modo teste)")

    return jsonify({
        'success': True,
        'token': token,
        'admin': username,
        'tipo_admin': admin_type,
        'turmas_permitidas': turmas_permitidas,
        'permissoes': permissoes,
        'is_global_admin': admin_type == 'global',
        'expires_in': JWT_EXPIRATION * 3600
    })
# ============================================
# MODO MANUTENÇÃO
# ============================================

MAINTENANCE_FILE = os.path.join(BASE_DIR, 'maintenance.json')

def get_maintenance_status():
    try:
        if os.path.exists(MAINTENANCE_FILE):
            with open(MAINTENANCE_FILE, 'r') as f:
                data = json.load(f)
                return data.get('maintenance_mode', False), data.get('message', '')
    except:
        pass
    return False, ''

def set_maintenance_status(enable, message=''):
    try:
        with open(MAINTENANCE_FILE, 'w') as f:
            json.dump({'maintenance_mode': enable, 'message': message, 'updated_at': datetime.now().isoformat(), 'updated_by': 'super_admin'}, f)
        return True
    except:
        return False

def maintenance_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if hasattr(request, 'user') and is_super_admin(request.user):
            return f(*args, **kwargs)
        is_maintenance, msg = get_maintenance_status()
        if is_maintenance:
            return jsonify({'error': 'Sistema em manutenção', 'message': msg or 'O sistema está passando por manutenção. Tente novamente mais tarde.', 'maintenance_mode': True}), 503
        return f(*args, **kwargs)
    return decorated

@app.route('/api/maintenance/status', methods=['GET'])
def maintenance_status():
    is_maintenance, message = get_maintenance_status()
    return jsonify({'maintenance_mode': is_maintenance, 'message': message})

@app.route('/api/maintenance/set', methods=['POST'])
@token_required
def maintenance_set():
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403

    data = request.get_json()
    enable = data.get('maintenance_mode', False)
    message = data.get('message', '')

    if set_maintenance_status(enable, message):
        action = 'ativou' if enable else 'desativou'
        log_admin_action_db(request.user, 'maintenance_mode', f'{action} o modo manutenção. Mensagem: {message}')
        return jsonify({'success': True, 'message': f'Modo manutenção {action} com sucesso', 'maintenance_mode': enable})
    else:
        return jsonify({'error': 'Erro ao alterar modo manutenção'}), 500

# ============================================================================
# AVISO DO SITE (POPUP DE ATUALIZAÇÃO) - GERENCIADO PELO SUPER ADMIN
# ============================================================================

SITE_NOTICE_FILE = os.path.join(BASE_DIR, 'site_notice.json')

def get_site_notice():
    """Retorna o aviso do site salvo em arquivo JSON"""
    try:
        if os.path.exists(SITE_NOTICE_FILE):
            with open(SITE_NOTICE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except:
        pass
    return {
        'active': False,
        'title': '',
        'body': '',
        'version_key': '',
        'start_at': None,
        'end_at': None,
        'updated_at': None,
        'updated_by': None
    }

def save_site_notice(data):
    """Salva o aviso do site em arquivo JSON"""
    try:
        with open(SITE_NOTICE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
        return True
    except:
        return False

@app.route('/api/public/site-notice', methods=['GET'])
def public_site_notice():
    """Retorna o aviso do site para o frontend público (sem auth)"""
    try:
        notice = get_site_notice()
        if not notice.get('active'):
            return jsonify({'active': False})

        # Verificar se está dentro do período de exibição
        now = datetime.now().isoformat()
        start_at = notice.get('start_at')
        end_at = notice.get('end_at')

        if start_at and now < start_at:
            return jsonify({'active': False})
        if end_at and now > end_at:
            return jsonify({'active': False})

        return jsonify({
            'active': True,
            'title': notice.get('title', ''),
            'body': notice.get('body', ''),
            'version_key': notice.get('version_key', ''),
            'start_at': start_at,
            'end_at': end_at
        })
    except Exception as e:
        print(f"Erro em public_site_notice: {e}")
        return jsonify({'active': False}), 200

@app.route('/api/super/site-notice', methods=['GET'])
@token_required
def super_get_site_notice():
    """Retorna o aviso do site completo para o Super Admin"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403
    try:
        notice = get_site_notice()
        return jsonify({'success': True, 'notice': notice})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/super/site-notice', methods=['POST'])
@token_required
def super_save_site_notice():
    """Salva o aviso do site (apenas Super Admin)"""
    if not is_super_admin(request.user):
        return jsonify({'error': 'Acesso negado. Apenas Super Admin.'}), 403
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Dados inválidos'}), 400

        title = data.get('title', '').strip()
        body = data.get('body', '').strip()
        active = bool(data.get('active', False))
        version_key = data.get('version_key', '').strip()
        start_at = data.get('start_at')
        end_at = data.get('end_at')

        if not title:
            return jsonify({'error': 'Título é obrigatório'}), 400
        if not body:
            return jsonify({'error': 'Conteúdo é obrigatório'}), 400
        if not version_key:
            return jsonify({'error': 'Chave de versão é obrigatória'}), 400

        notice = {
            'active': active,
            'title': title,
            'body': body,
            'version_key': version_key,
            'start_at': start_at,
            'end_at': end_at,
            'updated_at': datetime.now().isoformat(),
            'updated_by': request.user
        }

        if save_site_notice(notice):
            log_admin_action_db(request.user, 'site_notice', f'Aviso do site atualizado: {title} (ativo={active})')
            return jsonify({'success': True, 'message': 'Aviso do site salvo com sucesso', 'notice': notice})
        else:
            return jsonify({'error': 'Erro ao salvar aviso'}), 500

    except Exception as e:
        print(f"Erro em super_save_site_notice: {e}")
        return jsonify({'error': str(e)}), 500

# ============================================================================
# SISTEMA DE MENTORIA COMPLETO - VERSÃO 8.2 PRODUÇÃO
# COM CRUD COMPLETO E CORREÇÕES DE CORS
# ============================================================================

import secrets
import jwt
from datetime import datetime, timedelta
from functools import wraps
from flask import request, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
import time
import sqlite3

# ============================================================================
# CONSTANTES E CONFIGURAÇÕES DA MENTORIA
# ============================================================================

MENTOR_JWT_SECRET = secrets.token_hex(32)
MENTOR_JWT_EXPIRATION = 8

# ============================================================================
# FUNÇÕES AUXILIARES DE BANCO DE DADOS - MENTORIA
# ============================================================================

def converter_data_para_banco(data_str):
    """Converte string de data para formato do banco YYYY-MM-DD HH:MM:SS"""
    if not data_str:
        return None
    try:
        if ' ' in data_str and len(data_str) == 19:
            return data_str
        if 'T' in data_str:
            if '.' in data_str:
                data_str = data_str.split('.')[0]
            return data_str.replace('T', ' ')
        if len(data_str) == 10:
            return data_str + ' 00:00:00'
        return data_str
    except Exception as e:
        print(f"Erro ao converter data: {e}")
        return None

def execute_with_retry(func, max_retries=5, delay=0.1):
    """Executa uma função com retry em caso de database locked"""
    last_exception = None
    for attempt in range(max_retries):
        try:
            return func()
        except sqlite3.OperationalError as e:
            last_exception = e
            if 'database is locked' in str(e) and attempt < max_retries - 1:
                time.sleep(delay * (attempt + 1))
                continue
            raise e
        except Exception as e:
            raise e
    if last_exception:
        raise last_exception

def init_mentorship_db():
    """Cria todas as tabelas do sistema de mentoria"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_turmas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL,
            deleted_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_alunos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            turma_id INTEGER NOT NULL,
            deleted_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(nome, turma_id),
            FOREIGN KEY (turma_id) REFERENCES mentoria_turmas(id) ON DELETE CASCADE
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_mentores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            edv TEXT UNIQUE NOT NULL,
            senha_hash TEXT NOT NULL,
            ativo INTEGER DEFAULT 1,
            deleted_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_relacao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id INTEGER UNIQUE NOT NULL,
            mentor_id INTEGER NOT NULL,
            deleted_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (aluno_id) REFERENCES mentoria_alunos(id) ON DELETE CASCADE,
            FOREIGN KEY (mentor_id) REFERENCES mentoria_mentores(id) ON DELETE CASCADE
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_formularios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL,
            descricao TEXT,
            ativo INTEGER DEFAULT 1,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            deleted_at TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_perguntas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            formulario_id INTEGER NOT NULL,
            pergunta TEXT NOT NULL,
            tipo TEXT NOT NULL,
            obrigatoria INTEGER DEFAULT 1,
            ordem INTEGER NOT NULL,
            deleted_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (formulario_id) REFERENCES mentoria_formularios(id) ON DELETE CASCADE
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_opcoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pergunta_id INTEGER NOT NULL,
            opcao TEXT NOT NULL,
            ordem INTEGER DEFAULT 0,
            FOREIGN KEY (pergunta_id) REFERENCES mentoria_perguntas(id) ON DELETE CASCADE
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_ciclos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            form_id INTEGER,
            data_inicio TIMESTAMP NOT NULL,
            data_fim TIMESTAMP NOT NULL,
            ativo INTEGER DEFAULT 0,
            deleted_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_by TEXT,
            FOREIGN KEY (form_id) REFERENCES mentoria_formularios(id) ON DELETE SET NULL
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_respostas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id INTEGER NOT NULL,
            mentor_id INTEGER NOT NULL,
            ciclo_id INTEGER NOT NULL,
            data_envio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (aluno_id) REFERENCES mentoria_alunos(id) ON DELETE CASCADE,
            FOREIGN KEY (mentor_id) REFERENCES mentoria_mentores(id) ON DELETE CASCADE,
            FOREIGN KEY (ciclo_id) REFERENCES mentoria_ciclos(id) ON DELETE CASCADE,
            UNIQUE(aluno_id, ciclo_id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mentoria_respostas_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            resposta_id INTEGER NOT NULL,
            pergunta_id INTEGER NOT NULL,
            resposta TEXT NOT NULL,
            FOREIGN KEY (resposta_id) REFERENCES mentoria_respostas(id) ON DELETE CASCADE,
            FOREIGN KEY (pergunta_id) REFERENCES mentoria_perguntas(id) ON DELETE CASCADE
        )
    ''')

    cursor.execute('CREATE INDEX IF NOT EXISTS idx_respostas_aluno_ciclo ON mentoria_respostas(aluno_id, ciclo_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_respostas_itens_resposta ON mentoria_respostas_itens(resposta_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_perguntas_formulario ON mentoria_perguntas(formulario_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_opcoes_pergunta ON mentoria_opcoes(pergunta_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_ciclos_form ON mentoria_ciclos(form_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_ciclos_ativo ON mentoria_ciclos(ativo)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_perguntas_deleted ON mentoria_perguntas(deleted_at)')

    conn.commit()
    conn.close()
    print("✅ Tabelas de mentoria inicializadas")

def criar_formulario_padrao():
    """Cria um formulário padrão se não existir nenhum"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM mentoria_formularios WHERE deleted_at IS NULL")
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO mentoria_formularios (titulo, descricao, created_by)
            VALUES (?, ?, ?)
        ''', ('Formulário de Avaliação Socioemocional', 'Avaliação do desenvolvimento socioemocional do aluno na mentoria.', 'sistema'))
        form_id = cursor.lastrowid

        # 5 PERGUNTAS SOCIOEMOCIONAIS COM ESCALA 0-5
        perguntas_padrao = [
            ('Como você avalia a autoconfiança e autoestima do aluno atualmente?', 'multiple', 1, 1),
            ('Como você avalia a capacidade de comunicação e expressão do aluno?', 'multiple', 1, 2),
            ('Como você avalia o engajamento e comprometimento do aluno com a mentoria?', 'multiple', 1, 3),
            ('Como você avalia a resiliência do aluno diante de desafios e frustrações?', 'multiple', 1, 4),
            ('Como você avalia a autonomia e proatividade do aluno nas atividades?', 'multiple', 1, 5),
        ]

        for pergunta, tipo, obrigatoria, ordem in perguntas_padrao:
            cursor.execute('''
                INSERT INTO mentoria_perguntas (formulario_id, pergunta, tipo, obrigatoria, ordem)
                VALUES (?, ?, ?, ?, ?)
            ''', (form_id, pergunta, tipo, obrigatoria, ordem))

            if tipo == 'multiple':
                pergunta_id = cursor.lastrowid
                # Opções de 0 a 5
                opcoes = ['0 - Muito Insatisfatório', '1 - Insatisfatório', '2 - Regular', '3 - Bom', '4 - Muito Bom', '5 - Excelente']
                for idx, opt in enumerate(opcoes):
                    cursor.execute('INSERT INTO mentoria_opcoes (pergunta_id, opcao, ordem) VALUES (?, ?, ?)', (pergunta_id, opt, idx))

        print("✅ Formulário padrão com 5 perguntas socioemocionais (escala 0-5) criado")

    conn.commit()
    conn.close()

# ============================================================================
# FUNÇÕES DE CONSULTA E CRUD - MENTORIA
# ============================================================================

def get_active_cycle():
    """Retorna o ciclo ativo atual com seu formulário"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('''
                SELECT c.*, f.titulo as formulario_titulo, f.descricao as formulario_descricao, f.id as formulario_id
                FROM mentoria_ciclos c
                LEFT JOIN mentoria_formularios f ON c.form_id = f.id AND f.deleted_at IS NULL
                WHERE c.ativo = 1 AND c.deleted_at IS NULL
                AND c.data_inicio <= ? AND c.data_fim >= ?
                ORDER BY c.id DESC LIMIT 1
            ''', (now, now))
            result = cursor.fetchone()
            conn.close()
            return dict(result) if result else None
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em get_active_cycle: {e}")
        return None

def get_formulario_by_id(formulario_id):
    """Retorna um formulário completo com suas perguntas"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT * FROM mentoria_formularios WHERE id = ? AND deleted_at IS NULL', (formulario_id,))
            formulario = cursor.fetchone()

            if not formulario:
                conn.close()
                return None

            cursor.execute('SELECT * FROM mentoria_perguntas WHERE formulario_id = ? AND deleted_at IS NULL ORDER BY ordem', (formulario_id,))
            perguntas = cursor.fetchall()

            resultado = dict(formulario)
            resultado['perguntas'] = []

            for pergunta in perguntas:
                p = dict(pergunta)
                if p['tipo'] == 'multiple':
                    cursor.execute('SELECT opcao FROM mentoria_opcoes WHERE pergunta_id = ? ORDER BY ordem', (pergunta['id'],))
                    p['opcoes'] = [row['opcao'] for row in cursor.fetchall()]
                resultado['perguntas'].append(p)

            conn.close()
            return resultado
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em get_formulario_by_id: {e}")
        return None

def get_all_formularios():
    """Retorna todos os formulários ativos"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT f.*, COUNT(p.id) as total_perguntas
                FROM mentoria_formularios f
                LEFT JOIN mentoria_perguntas p ON f.id = p.formulario_id AND p.deleted_at IS NULL
                WHERE f.deleted_at IS NULL
                GROUP BY f.id
                ORDER BY f.created_at DESC
            ''')
            result = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return result
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em get_all_formularios: {e}")
        return []

def get_all_ciclos():
    """Retorna todos os ciclos"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT c.*, f.titulo as formulario_titulo,
                       COUNT(r.id) as total_respostas
                FROM mentoria_ciclos c
                LEFT JOIN mentoria_formularios f ON c.form_id = f.id AND f.deleted_at IS NULL
                LEFT JOIN mentoria_respostas r ON c.id = r.ciclo_id
                WHERE c.deleted_at IS NULL
                GROUP BY c.id
                ORDER BY c.created_at DESC
            ''')
            result = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return result
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em get_all_ciclos: {e}")
        return []

def get_aluno_by_id(aluno_id):
    """Busca aluno por ID"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT ma.*, mt.nome as turma_nome
                FROM mentoria_alunos ma
                JOIN mentoria_turmas mt ON ma.turma_id = mt.id
                WHERE ma.id = ? AND ma.deleted_at IS NULL
            ''', (aluno_id,))
            result = cursor.fetchone()
            conn.close()
            return dict(result) if result else None
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em get_aluno_by_id: {e}")
        return None

def get_mentor_by_id(mentor_id):
    """Busca mentor por ID"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT * FROM mentoria_mentores WHERE id = ? AND deleted_at IS NULL
            ''', (mentor_id,))
            result = cursor.fetchone()
            conn.close()
            return dict(result) if result else None
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em get_mentor_by_id: {e}")
        return None

def aluno_ja_respondeu_ciclo(aluno_id, ciclo_id):
    """Verifica se um aluno já respondeu um ciclo específico"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT id FROM mentoria_respostas WHERE aluno_id = ? AND ciclo_id = ?', (aluno_id, ciclo_id))
            result = cursor.fetchone()
            conn.close()
            return result is not None
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em aluno_ja_respondeu_ciclo: {e}")
        return False

def get_mentor_by_edv(edv):
    """Busca mentor por EDV"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT id, nome, edv, senha_hash, ativo FROM mentoria_mentores WHERE edv = ? AND deleted_at IS NULL', (edv,))
            result = cursor.fetchone()
            conn.close()
            return dict(result) if result else None
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em get_mentor_by_edv: {e}")
        return None

def get_alunos_por_mentor(mentor_id):
    """Retorna os alunos vinculados a um mentor"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT ma.id, ma.nome, mt.nome as turma_nome
                FROM mentoria_alunos ma
                JOIN mentoria_turmas mt ON ma.turma_id = mt.id
                JOIN mentoria_relacao mr ON ma.id = mr.aluno_id
                WHERE mr.mentor_id = ? AND mr.deleted_at IS NULL
                AND ma.deleted_at IS NULL AND mt.deleted_at IS NULL
            ''', (mentor_id,))
            result = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return result
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em get_alunos_por_mentor: {e}")
        return []

def get_respostas_aluno_por_ciclo(aluno_id):
    """Retorna todas as respostas de um aluno por ciclo para análise de evolução"""
    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT c.id as ciclo_id, c.nome as ciclo_nome, c.data_inicio, c.data_fim,
                       r.data_envio, ri.pergunta_id, ri.resposta, p.pergunta, p.tipo
                FROM mentoria_respostas r
                JOIN mentoria_ciclos c ON r.ciclo_id = c.id
                JOIN mentoria_respostas_itens ri ON r.id = ri.resposta_id
                JOIN mentoria_perguntas p ON ri.pergunta_id = p.id
                WHERE r.aluno_id = ? AND c.deleted_at IS NULL
                ORDER BY c.data_inicio, ri.pergunta_id
            ''', (aluno_id,))
            result = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return result
        return execute_with_retry(_query)
    except Exception as e:
        print(f"Erro em get_respostas_aluno_por_ciclo: {e}")
        return []

def salvar_respostas_mentor(aluno_id, mentor_id, ciclo_id, respostas):
    """Salva as respostas de um mentor para um aluno em um ciclo"""
    try:
        def _save():
            conn = get_db()
            cursor = conn.cursor()

            if aluno_ja_respondeu_ciclo(aluno_id, ciclo_id):
                conn.close()
                return False, "Aluno já respondeu este ciclo"

            cursor.execute('''
                INSERT INTO mentoria_respostas (aluno_id, mentor_id, ciclo_id, data_envio)
                VALUES (?, ?, ?, ?)
            ''', (aluno_id, mentor_id, ciclo_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            resposta_id = cursor.lastrowid

            for pergunta_id, resposta_texto in respostas.items():
                cursor.execute('''
                    INSERT INTO mentoria_respostas_itens (resposta_id, pergunta_id, resposta)
                    VALUES (?, ?, ?)
                ''', (resposta_id, pergunta_id, resposta_texto))

            conn.commit()
            conn.close()
            return True, "Respostas salvas com sucesso"
        return execute_with_retry(_save)
    except Exception as e:
        print(f"Erro em salvar_respostas_mentor: {e}")
        return False, f"Erro: {str(e)}"

def desativar_outros_ciclos(ciclo_id):
    """Desativa todos os outros ciclos ativos"""
    try:
        def _update():
            conn = get_db()
            cursor = conn.cursor()
            if ciclo_id:
                cursor.execute('UPDATE mentoria_ciclos SET ativo = 0 WHERE id != ? AND ativo = 1', (ciclo_id,))
            else:
                cursor.execute('UPDATE mentoria_ciclos SET ativo = 0 WHERE ativo = 1')
            conn.commit()
            conn.close()
            return True
        return execute_with_retry(_update)
    except Exception as e:
        print(f"Erro em desativar_outros_ciclos: {e}")
        return False

# ============================================================================
# WRAPERS DE AUTENTICAÇÃO - COM CORREÇÃO PARA PREFLIGHT REQUESTS
# ============================================================================

def mentoria_auth_required(f):
    """Wrapper para autenticação nas rotas de mentoria - APENAS ADMIN GLOBAL
       CORREÇÃO: Permite requisições OPTIONS (preflight CORS) passarem sem autenticação
    """
    @wraps(f)
    def decorated(*args, **kwargs):
        # 🔥 IMPORTANTE: Deixa requisições OPTIONS (preflight CORS) passarem sem autenticação
        if request.method == 'OPTIONS':
            return f(*args, **kwargs)

        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'error': True, 'message': 'Token não fornecido'}), 401

        token = auth_header.split(' ')[1]
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
            username = payload.get('username')
        except Exception as e:
            return jsonify({'error': True, 'message': f'Token inválido: {str(e)}'}), 401

        admin = get_admin_by_username(username)
        if not admin or admin['admin_type'] != 'global':
            return jsonify({'error': True, 'message': 'Acesso negado - Apenas Administrador Global'}), 403

        request.user = username
        request.admin_type = admin['admin_type']
        return f(*args, **kwargs)
    return decorated

def mentoria_mentor_auth_required(f):
    """Wrapper para autenticação de mentores
       CORREÇÃO: Permite requisições OPTIONS (preflight CORS) passarem sem autenticação
    """
    @wraps(f)
    def decorated(*args, **kwargs):
        # 🔥 IMPORTANTE: Deixa requisições OPTIONS (preflight CORS) passarem sem autenticação
        if request.method == 'OPTIONS':
            return f(*args, **kwargs)

        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'error': True, 'message': 'Token não fornecido'}), 401

        token = auth_header.split(' ')[1]
        try:
            payload = jwt.decode(token, MENTOR_JWT_SECRET, algorithms=['HS256'])
            mentor_id = payload.get('mentor_id')
            mentor_edv = payload.get('edv')
        except Exception as e:
            return jsonify({'error': True, 'message': f'Token inválido: {str(e)}'}), 401

        mentor = get_mentor_by_edv(mentor_edv)
        if not mentor:
            return jsonify({'error': True, 'message': 'Mentor não encontrado'}), 401

        if not mentor.get('ativo', 1):
            return jsonify({'error': True, 'message': 'Mentor desativado'}), 403

        request.mentor = mentor
        request.mentor_id = mentor_id
        return f(*args, **kwargs)
    return decorated

# ============================================================================
# ROTAS PÚBLICAS - LOGIN MENTOR
# ============================================================================

@app.route('/api/mentoria/mentor/login', methods=['POST', 'OPTIONS'])
def mentor_login():
    """Login do mentor com EDV e senha"""
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        edv = data.get('edv', '').strip()
        senha = data.get('senha', '').strip()

        if not edv or not senha:
            return jsonify({'success': False, 'message': 'EDV e senha são obrigatórios'}), 400

        mentor = get_mentor_by_edv(edv)

        if not mentor or not check_password_hash(mentor['senha_hash'], senha):
            return jsonify({'success': False, 'message': 'Credenciais inválidas'}), 401

        if not mentor.get('ativo', 1):
            return jsonify({'success': False, 'message': 'Conta desativada. Contate o administrador.'}), 403

        payload = {
            'mentor_id': mentor['id'],
            'edv': mentor['edv'],
            'nome': mentor['nome'],
            'exp': datetime.utcnow() + timedelta(hours=MENTOR_JWT_EXPIRATION),
            'iat': datetime.utcnow()
        }
        token = jwt.encode(payload, MENTOR_JWT_SECRET, algorithm='HS256')

        return jsonify({
            'success': True,
            'token': token,
            'mentor': {
                'id': mentor['id'],
                'nome': mentor['nome'],
                'edv': mentor['edv']
            }
        })
    except Exception as e:
        print(f"Erro no login mentor: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================================
# ROTAS DO MENTOR (PAINEL)
# ============================================================================

@app.route('/api/mentoria/mentor/me', methods=['GET', 'OPTIONS'])
@mentoria_mentor_auth_required
def mentor_get_info():
    """Retorna informações do mentor logado"""
    if request.method == 'OPTIONS':
        return '', 200

    try:
        mentor = request.mentor
        alunos = get_alunos_por_mentor(mentor['id'])
        ciclo_ativo = get_active_cycle()

        for aluno in alunos:
            if ciclo_ativo:
                aluno['ja_respondeu'] = aluno_ja_respondeu_ciclo(aluno['id'], ciclo_ativo['id'])
                aluno['pode_responder'] = not aluno['ja_respondeu']
            else:
                aluno['ja_respondeu'] = False
                aluno['pode_responder'] = False

        return jsonify({
            'success': True,
            'mentor': mentor,
            'alunos': alunos,
            'ciclo_ativo': ciclo_ativo
        })
    except Exception as e:
        print(f"Erro em mentor_get_info: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/mentoria/mentor/formulario', methods=['GET', 'OPTIONS'])
@mentoria_mentor_auth_required
def mentor_get_formulario():
    """Retorna o formulário do ciclo ativo para o mentor responder"""
    if request.method == 'OPTIONS':
        return '', 200

    try:
        ciclo_ativo = get_active_cycle()
        if not ciclo_ativo:
            return jsonify({'success': False, 'message': 'Nenhum ciclo ativo no momento'}), 404

        if not ciclo_ativo.get('formulario_id'):
            return jsonify({'success': False, 'message': 'Ciclo ativo não possui formulário associado'}), 404

        formulario = get_formulario_by_id(ciclo_ativo['formulario_id'])
        if not formulario:
            return jsonify({'success': False, 'message': 'Formulário não encontrado'}), 404

        aluno_id = request.args.get('aluno_id')
        if aluno_id:
            if aluno_ja_respondeu_ciclo(int(aluno_id), ciclo_ativo['id']):
                return jsonify({'success': False, 'message': 'Você já respondeu este formulário para este aluno'}), 403

        return jsonify({
            'success': True,
            'ciclo': {
                'id': ciclo_ativo['id'],
                'nome': ciclo_ativo['nome'],
                'data_inicio': ciclo_ativo['data_inicio'],
                'data_fim': ciclo_ativo['data_fim']
            },
            'formulario': formulario
        })
    except Exception as e:
        print(f"Erro em mentor_get_formulario: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/mentoria/mentor/responder', methods=['POST', 'OPTIONS'])
@mentoria_mentor_auth_required
def mentor_salvar_respostas():
    """Salva as respostas do mentor para um aluno"""
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        aluno_id = data.get('aluno_id')
        respostas = data.get('respostas', {})

        if not aluno_id:
            return jsonify({'success': False, 'message': 'Aluno não informado'}), 400

        if not respostas:
            return jsonify({'success': False, 'message': 'Nenhuma resposta enviada'}), 400

        ciclo_ativo = get_active_cycle()
        if not ciclo_ativo:
            return jsonify({'success': False, 'message': 'Nenhum ciclo ativo no momento'}), 404

        if aluno_ja_respondeu_ciclo(aluno_id, ciclo_ativo['id']):
            return jsonify({'success': False, 'message': 'Você já respondeu este formulário para este aluno'}), 403

        sucesso, mensagem = salvar_respostas_mentor(aluno_id, request.mentor['id'], ciclo_ativo['id'], respostas)

        if sucesso:
            log_admin_action(request.mentor['nome'], 'Responder Mentoria',
                           f'Mentor respondeu formulário para aluno ID {aluno_id} no ciclo {ciclo_ativo["nome"]}')
            return jsonify({'success': True, 'message': mensagem})
        else:
            return jsonify({'success': False, 'message': mensagem}), 400

    except Exception as e:
        print(f"Erro em mentor_salvar_respostas: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ============================================================================
# ROTAS DE ADMIN GLOBAL - TURMAS (CRUD COMPLETO)
# ============================================================================

@app.route('/api/mentoria/turmas', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_turmas_get():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT id, nome, created_at FROM mentoria_turmas WHERE deleted_at IS NULL ORDER BY nome')
            result = [{'id': row['id'], 'nome': row['nome'], 'created_at': row['created_at']} for row in cursor.fetchall()]
            conn.close()
            return result
        turmas = execute_with_retry(_query)
        return jsonify({'success': True, 'turmas': turmas})
    except Exception as e:
        print(f"Erro em mentoria_turmas_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/turmas', methods=['POST', 'OPTIONS'])
@mentoria_auth_required
def mentoria_turmas_post():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': True, 'message': 'Dados não fornecidos'}), 400

        nome = data.get('nome', '').strip()
        if not nome:
            return jsonify({'error': True, 'message': 'Nome da turma é obrigatório'}), 400

        def _save():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM mentoria_turmas WHERE nome = ? AND deleted_at IS NULL', (nome,))
            if cursor.fetchone():
                conn.close()
                return None, 'Turma já existe'

            cursor.execute('INSERT INTO mentoria_turmas (nome) VALUES (?)', (nome,))
            turma_id = cursor.lastrowid
            conn.commit()
            conn.close()
            return turma_id, None

        result, error = execute_with_retry(_save)
        if error:
            return jsonify({'error': True, 'message': error}), 409

        log_admin_action(request.user, 'Criar Turma Mentoria', f'Criada turma: {nome}')
        return jsonify({'success': True, 'id': result, 'nome': nome, 'message': 'Turma criada com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_turmas_post: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/turmas/<int:turma_id>', methods=['DELETE', 'OPTIONS'])
@mentoria_auth_required
def mentoria_turmas_delete(turma_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _delete():
            conn = get_db()
            cursor = conn.cursor()

            # Verificar se a turma existe
            cursor.execute('SELECT nome FROM mentoria_turmas WHERE id = ?', (turma_id,))
            turma = cursor.fetchone()
            if not turma:
                conn.close()
                return None

            # 🔥 HARD DELETE - Remove completamente os registros
            # 1. Remove relacionamentos (alunos desta turma)
            cursor.execute('''
                DELETE FROM mentoria_relacao
                WHERE aluno_id IN (SELECT id FROM mentoria_alunos WHERE turma_id = ?)
            ''', (turma_id,))

            # 2. Remove os alunos da turma
            cursor.execute('DELETE FROM mentoria_alunos WHERE turma_id = ?', (turma_id,))

            # 3. Remove a turma
            cursor.execute('DELETE FROM mentoria_turmas WHERE id = ?', (turma_id,))

            conn.commit()
            conn.close()
            return turma['nome']

        turma_nome = execute_with_retry(_delete)
        if not turma_nome:
            return jsonify({'error': True, 'message': 'Turma não encontrada'}), 404

        log_admin_action(request.user, 'Deletar Turma Mentoria', f'Deletada turma: {turma_nome}')
        return jsonify({'success': True, 'message': 'Turma removida com sucesso (Hard Delete)'})

    except Exception as e:
        print(f"Erro em mentoria_turmas_delete: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ROTAS DE ADMIN GLOBAL - ALUNOS (CRUD COMPLETO)
# ============================================================================

@app.route('/api/mentoria/alunos', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_alunos_get():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT ma.id, ma.nome, ma.turma_id, mt.nome as turma_nome,
                       CASE WHEN mr.id IS NOT NULL AND mr.deleted_at IS NULL THEN 1 ELSE 0 END as tem_mentor,
                       mm.nome as mentor_nome, mm.edv as mentor_edv, mm.id as mentor_id,
                       CASE WHEN r.id IS NOT NULL THEN 1 ELSE 0 END as respondeu
                FROM mentoria_alunos ma
                JOIN mentoria_turmas mt ON ma.turma_id = mt.id
                LEFT JOIN mentoria_relacao mr ON ma.id = mr.aluno_id AND mr.deleted_at IS NULL
                LEFT JOIN mentoria_mentores mm ON mr.mentor_id = mm.id AND mm.deleted_at IS NULL
                LEFT JOIN mentoria_respostas r ON ma.id = r.aluno_id
                WHERE ma.deleted_at IS NULL AND mt.deleted_at IS NULL
                GROUP BY ma.id
                ORDER BY mt.nome, ma.nome
            ''')
            result = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return result
        alunos = execute_with_retry(_query)
        return jsonify({'success': True, 'alunos': alunos})
    except Exception as e:
        print(f"Erro em mentoria_alunos_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/alunos', methods=['POST', 'OPTIONS'])
@mentoria_auth_required
def mentoria_alunos_post():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        nome = data.get('nome', '').strip()
        turma_id = data.get('turma_id')

        if not nome or not turma_id:
            return jsonify({'error': True, 'message': 'Nome e turma_id são obrigatórios'}), 400

        def _save():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM mentoria_turmas WHERE id = ? AND deleted_at IS NULL', (turma_id,))
            if not cursor.fetchone():
                conn.close()
                return None, 'Turma não encontrada'

            cursor.execute('SELECT id FROM mentoria_alunos WHERE nome = ? AND turma_id = ? AND deleted_at IS NULL', (nome, turma_id))
            if cursor.fetchone():
                conn.close()
                return None, 'Aluno já existe nesta turma'

            cursor.execute('INSERT INTO mentoria_alunos (nome, turma_id) VALUES (?, ?)', (nome, turma_id))
            aluno_id = cursor.lastrowid
            conn.commit()
            conn.close()
            return aluno_id, None

        aluno_id, error = execute_with_retry(_save)
        if error:
            return jsonify({'error': True, 'message': error}), 400

        log_admin_action(request.user, 'Criar Aluno Mentoria', f'Criado aluno: {nome}')
        return jsonify({'success': True, 'id': aluno_id, 'message': 'Aluno criado com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_alunos_post: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/alunos/<int:aluno_id>', methods=['PUT', 'OPTIONS'])
@mentoria_auth_required
def mentoria_alunos_put(aluno_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        nome = data.get('nome')
        turma_id = data.get('turma_id')

        if not nome and not turma_id:
            return jsonify({'error': True, 'message': 'Nenhum dado para atualizar'}), 400

        def _update():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM mentoria_alunos WHERE id = ? AND deleted_at IS NULL', (aluno_id,))
            if not cursor.fetchone():
                conn.close()
                return False, 'Aluno não encontrado'

            if turma_id:
                cursor.execute('SELECT id FROM mentoria_turmas WHERE id = ? AND deleted_at IS NULL', (turma_id,))
                if not cursor.fetchone():
                    conn.close()
                    return False, 'Turma não encontrada'

            updates = []
            params = []
            if nome:
                updates.append('nome = ?')
                params.append(nome)
            if turma_id:
                updates.append('turma_id = ?')
                params.append(turma_id)

            if updates:
                params.append(aluno_id)
                cursor.execute(f'UPDATE mentoria_alunos SET {", ".join(updates)} WHERE id = ?', params)

            conn.commit()
            conn.close()
            return True, None

        success, error = execute_with_retry(_update)
        if not success:
            return jsonify({'error': True, 'message': error}), 404

        log_admin_action(request.user, 'Atualizar Aluno Mentoria', f'Atualizado aluno ID {aluno_id}')
        return jsonify({'success': True, 'message': 'Aluno atualizado com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_alunos_put: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/alunos/<int:aluno_id>', methods=['DELETE', 'OPTIONS'])
@mentoria_auth_required
def mentoria_alunos_delete(aluno_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _delete():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT nome FROM mentoria_alunos WHERE id = ? AND deleted_at IS NULL', (aluno_id,))
            aluno = cursor.fetchone()
            if not aluno:
                conn.close()
                return None

            cursor.execute('UPDATE mentoria_relacao SET deleted_at = CURRENT_TIMESTAMP WHERE aluno_id = ?', (aluno_id,))
            cursor.execute('UPDATE mentoria_alunos SET deleted_at = CURRENT_TIMESTAMP WHERE id = ?', (aluno_id,))

            conn.commit()
            conn.close()
            return aluno['nome']

        aluno_nome = execute_with_retry(_delete)
        if not aluno_nome:
            return jsonify({'error': True, 'message': 'Aluno não encontrado'}), 404

        log_admin_action(request.user, 'Deletar Aluno Mentoria', f'Deletado aluno: {aluno_nome}')
        return jsonify({'success': True, 'message': 'Aluno removido com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_alunos_delete: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ROTAS DE ADMIN GLOBAL - MENTORES (CRUD COMPLETO)
# ============================================================================

@app.route('/api/mentoria/mentores', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_mentores_get():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT m.id, m.nome, m.edv, m.ativo, m.created_at,
                       COUNT(DISTINCT mr.aluno_id) as total_alunos
                FROM mentoria_mentores m
                LEFT JOIN mentoria_relacao mr ON m.id = mr.mentor_id AND mr.deleted_at IS NULL
                WHERE m.deleted_at IS NULL
                GROUP BY m.id
                ORDER BY m.nome
            ''')
            result = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return result
        mentores = execute_with_retry(_query)
        return jsonify({'success': True, 'mentores': mentores})
    except Exception as e:
        print(f"Erro em mentoria_mentores_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/mentores', methods=['POST', 'OPTIONS'])
@mentoria_auth_required
def mentoria_mentores_post():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        nome = data.get('nome', '').strip()
        edv = data.get('edv', '').strip()
        senha = data.get('senha', '123456').strip()
        ativo = data.get('ativo', 1)

        if not nome or not edv:
            return jsonify({'error': True, 'message': 'Nome e EDV são obrigatórios'}), 400

        def _save():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM mentoria_mentores WHERE edv = ? AND deleted_at IS NULL', (edv,))
            if cursor.fetchone():
                conn.close()
                return None, 'EDV já cadastrado'

            senha_hash = generate_password_hash(senha)
            cursor.execute('''
                INSERT INTO mentoria_mentores (nome, edv, senha_hash, ativo)
                VALUES (?, ?, ?, ?)
            ''', (nome, edv, senha_hash, ativo))
            mentor_id = cursor.lastrowid
            conn.commit()
            conn.close()
            return mentor_id, None

        mentor_id, error = execute_with_retry(_save)
        if error:
            return jsonify({'error': True, 'message': error}), 400

        log_admin_action(request.user, 'Criar Mentor Mentoria', f'Criado mentor: {nome} (EDV: {edv})')
        return jsonify({'success': True, 'id': mentor_id, 'message': 'Mentor criado com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_mentores_post: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/mentores/<int:mentor_id>', methods=['PUT', 'OPTIONS'])
@mentoria_auth_required
def mentoria_mentores_put(mentor_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        nome = data.get('nome')
        senha = data.get('senha')
        ativo = data.get('ativo')

        if not nome and not senha and ativo is None:
            return jsonify({'error': True, 'message': 'Nenhum dado para atualizar'}), 400

        def _update():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM mentoria_mentores WHERE id = ? AND deleted_at IS NULL', (mentor_id,))
            if not cursor.fetchone():
                conn.close()
                return False, 'Mentor não encontrado'

            updates = []
            params = []
            if nome:
                updates.append('nome = ?')
                params.append(nome)
            if senha:
                senha_hash = generate_password_hash(senha)
                updates.append('senha_hash = ?')
                params.append(senha_hash)
            if ativo is not None:
                updates.append('ativo = ?')
                params.append(ativo)

            if updates:
                params.append(mentor_id)
                cursor.execute(f'UPDATE mentoria_mentores SET {", ".join(updates)} WHERE id = ?', params)

            conn.commit()
            conn.close()
            return True, None

        success, error = execute_with_retry(_update)
        if not success:
            return jsonify({'error': True, 'message': error}), 404

        log_admin_action(request.user, 'Atualizar Mentor Mentoria', f'Atualizado mentor ID {mentor_id}')
        return jsonify({'success': True, 'message': 'Mentor atualizado com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_mentores_put: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/mentores/<int:mentor_id>', methods=['DELETE', 'OPTIONS'])
@mentoria_auth_required
def mentoria_mentores_delete(mentor_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _delete():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT nome FROM mentoria_mentores WHERE id = ? AND deleted_at IS NULL', (mentor_id,))
            mentor = cursor.fetchone()
            if not mentor:
                conn.close()
                return None

            cursor.execute('UPDATE mentoria_relacao SET deleted_at = CURRENT_TIMESTAMP WHERE mentor_id = ?', (mentor_id,))
            cursor.execute('UPDATE mentoria_mentores SET deleted_at = CURRENT_TIMESTAMP WHERE id = ?', (mentor_id,))

            conn.commit()
            conn.close()
            return mentor['nome']

        mentor_nome = execute_with_retry(_delete)
        if not mentor_nome:
            return jsonify({'error': True, 'message': 'Mentor não encontrado'}), 404

        log_admin_action(request.user, 'Deletar Mentor Mentoria', f'Deletado mentor: {mentor_nome}')
        return jsonify({'success': True, 'message': 'Mentor removido com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_mentores_delete: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ROTAS DE ADMIN GLOBAL - RELACIONAMENTOS (CRUD COMPLETO)
# ============================================================================

@app.route('/api/mentoria/relacoes', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_relacoes_get():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT mr.id, mr.created_at, mr.updated_at,
                       ma.id as aluno_id, ma.nome as aluno_nome,
                       mt.id as turma_id, mt.nome as turma_nome,
                       mm.id as mentor_id, mm.nome as mentor_nome, mm.edv as mentor_edv
                FROM mentoria_relacao mr
                JOIN mentoria_alunos ma ON mr.aluno_id = ma.id
                JOIN mentoria_turmas mt ON ma.turma_id = mt.id
                JOIN mentoria_mentores mm ON mr.mentor_id = mm.id
                WHERE mr.deleted_at IS NULL AND ma.deleted_at IS NULL AND mt.deleted_at IS NULL AND mm.deleted_at IS NULL
                ORDER BY mt.nome, ma.nome
            ''')
            result = [dict(row) for row in cursor.fetchall()]
            conn.close()
            return result
        relacoes = execute_with_retry(_query)
        return jsonify({'success': True, 'relacoes': relacoes})
    except Exception as e:
        print(f"Erro em mentoria_relacoes_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/relacoes', methods=['POST', 'OPTIONS'])
@mentoria_auth_required
def mentoria_relacoes_post():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        aluno_id = data.get('aluno_id')
        mentor_id = data.get('mentor_id')

        if not aluno_id or not mentor_id:
            return jsonify({'error': True, 'message': 'Aluno e mentor são obrigatórios'}), 400

        def _save():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT nome FROM mentoria_alunos WHERE id = ? AND deleted_at IS NULL', (aluno_id,))
            if not cursor.fetchone():
                conn.close()
                return None, 'Aluno não encontrado'

            cursor.execute('SELECT nome FROM mentoria_mentores WHERE id = ? AND deleted_at IS NULL', (mentor_id,))
            if not cursor.fetchone():
                conn.close()
                return None, 'Mentor não encontrado'

            cursor.execute('SELECT id FROM mentoria_relacao WHERE aluno_id = ?', (aluno_id,))
            existing = cursor.fetchone()

            if existing:
                cursor.execute('''
                    UPDATE mentoria_relacao SET mentor_id = ?, updated_at = CURRENT_TIMESTAMP, deleted_at = NULL
                    WHERE aluno_id = ?
                ''', (mentor_id, aluno_id))
                mensagem = 'Relacionamento atualizado com sucesso'
            else:
                cursor.execute('''
                    INSERT INTO mentoria_relacao (aluno_id, mentor_id)
                    VALUES (?, ?)
                ''', (aluno_id, mentor_id))
                mensagem = 'Relacionamento criado com sucesso'

            conn.commit()
            conn.close()
            return mensagem, None

        mensagem, error = execute_with_retry(_save)
        if error:
            return jsonify({'error': True, 'message': error}), 404

        log_admin_action(request.user, 'Gerenciar Relacionamento Mentoria', f'Relacionamento criado/atualizado')
        return jsonify({'success': True, 'message': mensagem})
    except Exception as e:
        print(f"Erro em mentoria_relacoes_post: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/relacoes/<int:relacao_id>', methods=['DELETE', 'OPTIONS'])
@mentoria_auth_required
def mentoria_relacoes_delete(relacao_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _delete():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('UPDATE mentoria_relacao SET deleted_at = CURRENT_TIMESTAMP WHERE id = ?', (relacao_id,))
            conn.commit()
            conn.close()
            return True
        execute_with_retry(_delete)
        return jsonify({'success': True, 'message': 'Relação removida com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_relacoes_delete: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ROTAS DE ADMIN GLOBAL - FORMULÁRIOS (CRUD COMPLETO)
# ============================================================================

@app.route('/api/mentoria/formularios', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_formularios_get():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        formularios = get_all_formularios()
        return jsonify({'success': True, 'formularios': formularios})
    except Exception as e:
        print(f"Erro em mentoria_formularios_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/formularios', methods=['POST', 'OPTIONS'])
@mentoria_auth_required
def mentoria_formularios_post():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        titulo = data.get('titulo', '').strip()
        descricao = data.get('descricao', '').strip()
        perguntas = data.get('perguntas', [])

        if not titulo:
            return jsonify({'error': True, 'message': 'Título é obrigatório'}), 400

        if len(perguntas) == 0:
            return jsonify({'error': True, 'message': 'Adicione pelo menos uma pergunta'}), 400

        def _save():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('''
                INSERT INTO mentoria_formularios (titulo, descricao, created_by)
                VALUES (?, ?, ?)
            ''', (titulo, descricao, request.user))
            form_id = cursor.lastrowid

            for ordem, pergunta_data in enumerate(perguntas, 1):
                pergunta_texto = pergunta_data.get('pergunta', '').strip()
                tipo = pergunta_data.get('tipo', 'text')
                obrigatoria = 1 if pergunta_data.get('obrigatoria', True) else 0

                if not pergunta_texto:
                    continue

                cursor.execute('''
                    INSERT INTO mentoria_perguntas (formulario_id, pergunta, tipo, obrigatoria, ordem)
                    VALUES (?, ?, ?, ?, ?)
                ''', (form_id, pergunta_texto, tipo, obrigatoria, ordem))
                pergunta_id = cursor.lastrowid

                if tipo == 'multiple':
                    opcoes = pergunta_data.get('opcoes', [])
                    for opt_ordem, opcao in enumerate(opcoes):
                        if opcao.strip():
                            cursor.execute('''
                                INSERT INTO mentoria_opcoes (pergunta_id, opcao, ordem)
                                VALUES (?, ?, ?)
                            ''', (pergunta_id, opcao.strip(), opt_ordem))

            conn.commit()
            conn.close()
            return form_id

        form_id = execute_with_retry(_save)

        log_admin_action(request.user, 'Criar Formulário Mentoria', f'Criado formulário: {titulo}')
        return jsonify({'success': True, 'id': form_id, 'message': 'Formulário criado com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_formularios_post: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/formularios/<int:formulario_id>', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_formularios_detail_get(formulario_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        formulario = get_formulario_by_id(formulario_id)
        if not formulario:
            return jsonify({'error': True, 'message': 'Formulário não encontrado'}), 404
        return jsonify({'success': True, 'formulario': formulario})
    except Exception as e:
        print(f"Erro em mentoria_formularios_detail_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/formularios/<int:formulario_id>', methods=['PUT', 'OPTIONS'])
@mentoria_auth_required
def mentoria_formularios_put(formulario_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        titulo = data.get('titulo')
        descricao = data.get('descricao')
        ativo = data.get('ativo')

        if not titulo and not descricao and ativo is None:
            return jsonify({'error': True, 'message': 'Nenhum dado para atualizar'}), 400

        def _update():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM mentoria_formularios WHERE id = ? AND deleted_at IS NULL', (formulario_id,))
            if not cursor.fetchone():
                conn.close()
                return False, 'Formulário não encontrado'

            updates = []
            params = []
            if titulo:
                updates.append('titulo = ?')
                params.append(titulo)
            if descricao is not None:
                updates.append('descricao = ?')
                params.append(descricao)
            if ativo is not None:
                updates.append('ativo = ?')
                params.append(ativo)

            updates.append('updated_at = CURRENT_TIMESTAMP')
            params.append(formulario_id)

            cursor.execute(f'UPDATE mentoria_formularios SET {", ".join(updates)} WHERE id = ?', params)

            conn.commit()
            conn.close()
            return True, None

        success, error = execute_with_retry(_update)
        if not success:
            return jsonify({'error': True, 'message': error}), 404

        log_admin_action(request.user, 'Atualizar Formulário Mentoria', f'Atualizado formulário ID {formulario_id}')
        return jsonify({'success': True, 'message': 'Formulário atualizado com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_formularios_put: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/formularios/<int:formulario_id>', methods=['DELETE', 'OPTIONS'])
@mentoria_auth_required
def mentoria_formularios_delete(formulario_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _delete():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT titulo FROM mentoria_formularios WHERE id = ? AND deleted_at IS NULL', (formulario_id,))
            formulario = cursor.fetchone()
            if not formulario:
                conn.close()
                return None

            cursor.execute('UPDATE mentoria_formularios SET deleted_at = CURRENT_TIMESTAMP WHERE id = ?', (formulario_id,))
            cursor.execute('UPDATE mentoria_perguntas SET deleted_at = CURRENT_TIMESTAMP WHERE formulario_id = ?', (formulario_id,))

            conn.commit()
            conn.close()
            return formulario['titulo']

        titulo = execute_with_retry(_delete)
        if not titulo:
            return jsonify({'error': True, 'message': 'Formulário não encontrado'}), 404

        log_admin_action(request.user, 'Deletar Formulário Mentoria', f'Deletado formulário: {titulo}')
        return jsonify({'success': True, 'message': 'Formulário removido com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_formularios_delete: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ROTAS DE ADMIN GLOBAL - CICLOS (CRUD COMPLETO)
# ============================================================================

@app.route('/api/mentoria/ciclos', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_ciclos_get():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        ciclos = get_all_ciclos()
        return jsonify({'success': True, 'ciclos': ciclos})
    except Exception as e:
        print(f"Erro em mentoria_ciclos_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/ciclos', methods=['POST', 'OPTIONS'])
@mentoria_auth_required
def mentoria_ciclos_post():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        nome = data.get('nome', '').strip()
        form_id = data.get('form_id')
        data_inicio = converter_data_para_banco(data.get('data_inicio'))
        data_fim = converter_data_para_banco(data.get('data_fim'))
        ativo = data.get('ativo', False)

        if not nome:
            return jsonify({'error': True, 'message': 'Nome é obrigatório'}), 400

        if not form_id:
            return jsonify({'error': True, 'message': 'Selecione um formulário'}), 400

        if not data_inicio or not data_fim:
            return jsonify({'error': True, 'message': 'Datas são obrigatórias'}), 400

        def _save():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM mentoria_formularios WHERE id = ? AND deleted_at IS NULL', (form_id,))
            if not cursor.fetchone():
                conn.close()
                return None, 'Formulário não encontrado'

            if ativo:
                desativar_outros_ciclos(None)

            cursor.execute('''
                INSERT INTO mentoria_ciclos (nome, form_id, data_inicio, data_fim, ativo, created_by)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (nome, form_id, data_inicio, data_fim, 1 if ativo else 0, request.user))
            ciclo_id = cursor.lastrowid
            conn.commit()
            conn.close()
            return ciclo_id, None

        ciclo_id, error = execute_with_retry(_save)
        if error:
            return jsonify({'error': True, 'message': error}), 404

        log_admin_action(request.user, 'Criar Ciclo Mentoria', f'Criado ciclo: {nome}')
        return jsonify({'success': True, 'id': ciclo_id, 'message': 'Ciclo criado com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_ciclos_post: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/ciclos/<int:ciclo_id>', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_ciclos_detail_get(ciclo_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT c.*, f.titulo as formulario_titulo
                FROM mentoria_ciclos c
                LEFT JOIN mentoria_formularios f ON c.form_id = f.id AND f.deleted_at IS NULL
                WHERE c.id = ? AND c.deleted_at IS NULL
            ''', (ciclo_id,))
            result = cursor.fetchone()
            conn.close()
            return dict(result) if result else None
        ciclo = execute_with_retry(_query)
        if not ciclo:
            return jsonify({'error': True, 'message': 'Ciclo não encontrado'}), 404
        return jsonify({'success': True, 'ciclo': ciclo})
    except Exception as e:
        print(f"Erro em mentoria_ciclos_detail_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/ciclos/<int:ciclo_id>', methods=['PUT', 'OPTIONS'])
@mentoria_auth_required
def mentoria_ciclos_put(ciclo_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        nome = data.get('nome')
        form_id = data.get('form_id')
        data_inicio = converter_data_para_banco(data.get('data_inicio')) if data.get('data_inicio') else None
        data_fim = converter_data_para_banco(data.get('data_fim')) if data.get('data_fim') else None
        ativo = data.get('ativo')

        if not nome and not form_id and not data_inicio and not data_fim and ativo is None:
            return jsonify({'error': True, 'message': 'Nenhum dado para atualizar'}), 400

        def _update():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM mentoria_ciclos WHERE id = ? AND deleted_at IS NULL', (ciclo_id,))
            if not cursor.fetchone():
                conn.close()
                return False, 'Ciclo não encontrado'

            if form_id:
                cursor.execute('SELECT id FROM mentoria_formularios WHERE id = ? AND deleted_at IS NULL', (form_id,))
                if not cursor.fetchone():
                    conn.close()
                    return False, 'Formulário não encontrado'

            if ativo:
                desativar_outros_ciclos(ciclo_id)

            updates = []
            params = []
            if nome:
                updates.append('nome = ?')
                params.append(nome)
            if form_id:
                updates.append('form_id = ?')
                params.append(form_id)
            if data_inicio:
                updates.append('data_inicio = ?')
                params.append(data_inicio)
            if data_fim:
                updates.append('data_fim = ?')
                params.append(data_fim)
            if ativo is not None:
                updates.append('ativo = ?')
                params.append(1 if ativo else 0)

            if updates:
                params.append(ciclo_id)
                cursor.execute(f'UPDATE mentoria_ciclos SET {", ".join(updates)} WHERE id = ?', params)

            conn.commit()
            conn.close()
            return True, None

        success, error = execute_with_retry(_update)
        if not success:
            return jsonify({'error': True, 'message': error}), 404

        log_admin_action(request.user, 'Atualizar Ciclo Mentoria', f'Atualizado ciclo ID {ciclo_id}')
        return jsonify({'success': True, 'message': 'Ciclo atualizado com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_ciclos_put: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mentoria/ciclos/<int:ciclo_id>', methods=['DELETE', 'OPTIONS'])
@mentoria_auth_required
def mentoria_ciclos_delete(ciclo_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _delete():
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT nome FROM mentoria_ciclos WHERE id = ? AND deleted_at IS NULL', (ciclo_id,))
            ciclo = cursor.fetchone()
            if not ciclo:
                conn.close()
                return None

            cursor.execute('UPDATE mentoria_ciclos SET deleted_at = CURRENT_TIMESTAMP WHERE id = ?', (ciclo_id,))

            conn.commit()
            conn.close()
            return ciclo['nome']

        nome = execute_with_retry(_delete)
        if not nome:
            return jsonify({'error': True, 'message': 'Ciclo não encontrado'}), 404

        log_admin_action(request.user, 'Deletar Ciclo Mentoria', f'Deletado ciclo: {nome}')
        return jsonify({'success': True, 'message': 'Ciclo removido com sucesso'})
    except Exception as e:
        print(f"Erro em mentoria_ciclos_delete: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ROTAS DE ADMIN GLOBAL - STATUS E DASHBOARD
# ============================================================================

@app.route('/api/mentoria/status', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_status_get():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT COUNT(*) as total FROM mentoria_alunos WHERE deleted_at IS NULL')
            total_alunos = cursor.fetchone()['total'] or 0

            cursor.execute('SELECT COUNT(*) as total FROM mentoria_relacao WHERE deleted_at IS NULL')
            alunos_com_mentor = cursor.fetchone()['total'] or 0

            cursor.execute('SELECT COUNT(*) as total FROM mentoria_mentores WHERE ativo = 1 AND deleted_at IS NULL')
            total_mentores = cursor.fetchone()['total'] or 0

            cursor.execute('SELECT COUNT(*) as total FROM mentoria_formularios WHERE deleted_at IS NULL')
            total_formularios = cursor.fetchone()['total'] or 0

            cursor.execute('SELECT COUNT(*) as total FROM mentoria_ciclos WHERE deleted_at IS NULL')
            total_ciclos = cursor.fetchone()['total'] or 0

            conn.close()
            return {
                'total_alunos': total_alunos,
                'alunos_com_mentor': alunos_com_mentor,
                'total_mentores': total_mentores,
                'total_formularios': total_formularios,
                'total_ciclos': total_ciclos
            }

        resumo = execute_with_retry(_query)
        active_cycle = get_active_cycle()
        ciclo_info = None

        if active_cycle:
            def _count():
                conn = get_db()
                cursor = conn.cursor()
                cursor.execute('SELECT COUNT(*) as total FROM mentoria_respostas WHERE ciclo_id = ?', (active_cycle['id'],))
                total = cursor.fetchone()['total'] or 0
                conn.close()
                return total

            total_submissoes = execute_with_retry(_count)
            percentual = round((total_submissoes / resumo['total_alunos']) * 100, 1) if resumo['total_alunos'] > 0 else 0
            ciclo_info = dict(active_cycle)
            ciclo_info['total_submissoes'] = total_submissoes
            ciclo_info['percentual_submissoes'] = percentual

        return jsonify({
            'success': True,
            'resumo': {
                'total_alunos': resumo['total_alunos'],
                'alunos_com_mentor': resumo['alunos_com_mentor'],
                'alunos_sem_mentor': max(0, resumo['total_alunos'] - resumo['alunos_com_mentor']),
                'total_mentores': resumo['total_mentores'],
                'total_formularios': resumo['total_formularios'],
                'total_ciclos': resumo['total_ciclos']
            },
            'ciclo_ativo': ciclo_info
        })
    except Exception as e:
        print(f"Erro em mentoria_status_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ROTAS DE ADMIN GLOBAL - RESPOSTAS
# ============================================================================

@app.route('/api/mentoria/respostas/aluno/<int:aluno_id>', methods=['GET', 'OPTIONS'])
@mentoria_auth_required
def mentoria_respostas_aluno_get(aluno_id):
    if request.method == 'OPTIONS':
        return '', 200

    try:
        def _query():
            conn = get_db()
            cursor = conn.cursor()

            cursor.execute('SELECT nome FROM mentoria_alunos WHERE id = ? AND deleted_at IS NULL', (aluno_id,))
            if not cursor.fetchone():
                conn.close()
                return None

            cursor.execute('''
                SELECT c.id as ciclo_id, c.nome as ciclo_nome, c.data_inicio, c.data_fim,
                       r.data_envio, mm.nome as mentor_nome, mm.edv as mentor_edv,
                       ri.pergunta_id, ri.resposta, p.pergunta, p.tipo
                FROM mentoria_respostas r
                JOIN mentoria_ciclos c ON r.ciclo_id = c.id
                JOIN mentoria_mentores mm ON r.mentor_id = mm.id
                JOIN mentoria_respostas_itens ri ON r.id = ri.resposta_id
                JOIN mentoria_perguntas p ON ri.pergunta_id = p.id
                WHERE r.aluno_id = ? AND c.deleted_at IS NULL
                ORDER BY c.data_inicio DESC, ri.pergunta_id
            ''', (aluno_id,))

            resultados = cursor.fetchall()
            conn.close()
            return resultados

        resultados = execute_with_retry(_query)
        if resultados is None:
            return jsonify({'error': True, 'message': 'Aluno não encontrado'}), 404

        ciclos_dict = {}
        for item in resultados:
            ciclo_id = item['ciclo_id']
            if ciclo_id not in ciclos_dict:
                ciclos_dict[ciclo_id] = {
                    'id': ciclo_id,
                    'nome': item['ciclo_nome'],
                    'data_inicio': item['data_inicio'],
                    'data_fim': item['data_fim'],
                    'data_envio': item['data_envio'],
                    'mentor': {
                        'nome': item['mentor_nome'],
                        'edv': item['mentor_edv']
                    },
                    'respostas': []
                }
            ciclos_dict[ciclo_id]['respostas'].append({
                'pergunta': item['pergunta'],
                'resposta': item['resposta'],
                'tipo': item['tipo']
            })

        return jsonify({
            'success': True,
            'respostas': list(ciclos_dict.values())
        })
    except Exception as e:
        print(f"Erro em mentoria_respostas_aluno_get: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# ROTAS DE ADMIN GLOBAL - IMPORTAÇÃO EM LOTE
# ============================================================================

@app.route('/api/mentoria/importar', methods=['POST', 'OPTIONS'])
@mentoria_auth_required
def mentoria_importar_completo():
    if request.method == 'OPTIONS':
        return '', 200

    try:
        data = request.get_json()
        if not isinstance(data, list):
            return jsonify({'error': True, 'message': 'Dados inválidos. Envie um array de objetos'}), 400

        if len(data) == 0:
            return jsonify({'error': True, 'message': 'Nenhum dado para importar'}), 400

        def _import():
            conn = get_db()
            cursor = conn.cursor()

            importados = 0
            atualizados = 0
            erros = []

            for idx, item in enumerate(data, 1):
                turma_nome = item.get('turma_nome', '').strip()
                aluno_nome = item.get('aluno_nome', '').strip()
                mentor_nome = item.get('mentor_nome', '').strip()
                mentor_edv = item.get('mentor_edv', '').strip()
                mentor_senha = item.get('mentor_senha', '123456').strip()

                if not turma_nome or not aluno_nome or not mentor_nome or not mentor_edv:
                    erros.append(f"Linha {idx}: Campos obrigatórios ausentes")
                    continue

                # 1. Verificar/Criar Turma
                cursor.execute('SELECT id FROM mentoria_turmas WHERE nome = ? AND deleted_at IS NULL', (turma_nome,))
                turma = cursor.fetchone()

                if not turma:
                    cursor.execute('INSERT INTO mentoria_turmas (nome) VALUES (?)', (turma_nome,))
                    turma_id = cursor.lastrowid
                else:
                    turma_id = turma['id']

                # 2. Verificar/Criar Aluno
                cursor.execute('SELECT id FROM mentoria_alunos WHERE nome = ? AND turma_id = ? AND deleted_at IS NULL', (aluno_nome, turma_id))
                aluno = cursor.fetchone()

                if not aluno:
                    cursor.execute('INSERT INTO mentoria_alunos (nome, turma_id) VALUES (?, ?)', (aluno_nome, turma_id))
                    aluno_id = cursor.lastrowid
                    importados += 1
                else:
                    aluno_id = aluno['id']

                # 3. Verificar/Criar Mentor
                cursor.execute('SELECT id FROM mentoria_mentores WHERE edv = ? AND deleted_at IS NULL', (mentor_edv,))
                mentor = cursor.fetchone()

                senha_hash = generate_password_hash(mentor_senha)

                if not mentor:
                    cursor.execute('''
                        INSERT INTO mentoria_mentores (nome, edv, senha_hash, ativo)
                        VALUES (?, ?, ?, 1)
                    ''', (mentor_nome, mentor_edv, senha_hash))
                    mentor_id = cursor.lastrowid
                    importados += 1
                else:
                    mentor_id = mentor['id']
                    cursor.execute('UPDATE mentoria_mentores SET nome = ?, senha_hash = ? WHERE id = ?', (mentor_nome, senha_hash, mentor_id))

                # 4. Criar/Atualizar Relacionamento
                cursor.execute('SELECT id FROM mentoria_relacao WHERE aluno_id = ? AND deleted_at IS NULL', (aluno_id,))
                relacao = cursor.fetchone()

                if relacao:
                    cursor.execute('UPDATE mentoria_relacao SET mentor_id = ?, deleted_at = NULL, updated_at = CURRENT_TIMESTAMP WHERE aluno_id = ?', (mentor_id, aluno_id))
                    atualizados += 1
                else:
                    cursor.execute('INSERT INTO mentoria_relacao (aluno_id, mentor_id) VALUES (?, ?)', (aluno_id, mentor_id))
                    importados += 1

            conn.commit()
            conn.close()
            return importados, atualizados, erros

        importados, atualizados, erros = execute_with_retry(_import)

        log_admin_action(request.user, 'Importação Completa Mentoria',
                        f'Importados: {importados}, Atualizados: {atualizados}, Erros: {len(erros)}')

        return jsonify({
            'success': True,
            'importados': importados,
            'atualizados': atualizados,
            'erros': erros[:20]
        })
    except Exception as e:
        print(f"Erro em mentoria_importar_completo: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# INICIALIZAÇÃO DO MÓDULO DE MENTORIA
# ============================================================================

def init_mentorship_module():
    try:
        init_mentorship_db()
        criar_formulario_padrao()
        print("=" * 50)
        print("✅ MÓDULO DE MENTORIA v8.2 INICIALIZADO COM SUCESSO!")
        print("=" * 50)
        print("📋 TURMAS: GET, POST, DELETE /api/mentoria/turmas")
        print("👥 ALUNOS: GET, POST, PUT, DELETE /api/mentoria/alunos")
        print("👨‍🏫 MENTORES: GET, POST, PUT, DELETE /api/mentoria/mentores")
        print("🔗 RELACIONAMENTOS: GET, POST, DELETE /api/mentoria/relacoes")
        print("📝 FORMULÁRIOS: GET, POST, PUT, DELETE /api/mentoria/formularios")
        print("🔄 CICLOS: GET, POST, PUT, DELETE /api/mentoria/ciclos")
        print("📊 STATUS: GET /api/mentoria/status")
        print("📥 IMPORTAÇÃO: POST /api/mentoria/importar")
        print("👨‍🎓 LOGIN MENTOR: POST /api/mentoria/mentor/login")
        print("✅ CRUD completo implementado para todas as entidades!")
        print("✅ Correção CORS para preflight requests implementada!")
        print("=" * 50)
    except Exception as e:
        print(f"⚠️ Erro ao inicializar módulo de mentoria: {e}")

# ============================================================================
# INICIALIZAR MÓDULO DE MENTORIA NA INICIALIZAÇÃO DO APP
# ============================================================================

def init_app():
    init_db()
    init_admins_db()
    migrate_presenca_table()

    # INICIALIZAR MÓDULOS DE MENTORIA
    init_mentorship_module()

    # Criar super admin padrão se não existir
    conn = get_admins_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM admins")
    count = cursor.fetchone()[0]
    conn.close()

    if count == 0:
        create_admin('erickdev', '030680901Erick$', 'global', created_by='Sistema')
        print("✅ Super admin padrão criado: erickdev / 030680901Erick$")

    print("=" * 50)
    print("✅ SISTEMA DE ESCALA DE LIMPEZA - VERSÃO  COMPLETO")



if __name__ == '__main__':
    init_app()
    app.run(debug=False, host='0.0.0.0', port=5002)
